#!/usr/bin/env python3
"""Generate ChineseEEG embeddings for the prespecified NeuroSem model-family panel.

The model family and representation rule are fixed by --model-key. The script resolves
and records the exact Hugging Face commit SHA before loading the model so each run is
reproducible even when the repository's main branch changes later.

This script is for LittlePrince runs 01-06 only during model screening. Run 07 should
remain untouched until model selection is frozen.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

MODEL_SPECS = {
    "xlm-roberta-large": {
        "model_id": "FacebookAI/xlm-roberta-large",
        "pooling": "mean_non_special",
        "prefix": "",
        "normalize": True,
    },
    "multilingual-e5-large": {
        "model_id": "intfloat/multilingual-e5-large",
        "pooling": "mean_attention",
        "prefix": "query: ",
        "normalize": True,
    },
    "bge-m3": {
        "model_id": "BAAI/bge-m3",
        "pooling": "cls",
        "prefix": "",
        "normalize": True,
    },
    "qwen3-embedding-0.6b": {
        "model_id": "Qwen/Qwen3-Embedding-0.6B",
        "pooling": "last_token",
        "prefix": "",
        "normalize": True,
    },
}


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[str] = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    denom = mask.sum(dim=1).clamp_min(1.0)
    return (hidden * mask).sum(dim=1) / denom


def l2_normalize(x):
    import torch
    return torch.nn.functional.normalize(x, p=2, dim=1)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate prespecified model-panel embeddings for ChineseEEG.")
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--run-number", type=int, required=True)
    parser.add_argument("--model-key", choices=sorted(MODEL_SPECS), required=True)
    parser.add_argument("--revision", default=None, help="Optional exact HF revision; otherwise resolve current repository SHA once and pin it.")
    parser.add_argument("--batch-size", type=int, default=16)
    parser.add_argument("--max-length", type=int, default=64)
    parser.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/chineseeeg_model_panel_embeddings"))
    args = parser.parse_args()

    if args.run_number == 7:
        raise SystemExit("Run 07 is sealed for cross-model holdout. Do not generate panel embeddings before model selection is frozen.")
    if args.run_number < 1 or args.run_number > 6:
        raise SystemExit("Model-panel screening is currently restricted to runs 01-06")

    try:
        import torch
        import transformers
        from huggingface_hub import HfApi
        from transformers import AutoModel, AutoTokenizer
    except ImportError as exc:
        raise SystemExit("Missing embedding dependencies. Install with: pip install -r requirements-embeddings.txt") from exc

    spec = MODEL_SPECS[args.model_key]
    model_id = spec["model_id"]
    revision = args.revision or HfApi().model_info(model_id).sha
    if not revision:
        raise SystemExit(f"Could not resolve exact revision for {model_id}")

    dataset = args.dataset.expanduser().resolve()
    workbook = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / f"segmented_Chinense_novel_run_{args.run_number}.xlsx"
    if not workbook.exists():
        raise SystemExit(f"Workbook not materialized: {workbook}")
    texts = read_rows(workbook)
    if not texts:
        raise SystemExit("No workbook rows found")
    model_texts = [spec["prefix"] + t for t in texts]

    if args.device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    else:
        device = args.device
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("--device cuda requested but CUDA is unavailable")

    tokenizer = AutoTokenizer.from_pretrained(model_id, revision=revision, trust_remote_code=True)
    model = AutoModel.from_pretrained(model_id, revision=revision, trust_remote_code=True)
    model.eval().to(device)

    chunks = []
    token_counts: list[int] = []
    truncated_rows = 0

    with torch.inference_mode():
        for start in range(0, len(model_texts), args.batch_size):
            batch = model_texts[start:start + args.batch_size]
            encoded = tokenizer(
                batch,
                padding=True,
                truncation=True,
                max_length=args.max_length,
                return_tensors="pt",
                return_special_tokens_mask=True,
            )
            special_mask = encoded.pop("special_tokens_mask")
            attention = encoded["attention_mask"]
            token_counts.extend(attention.sum(dim=1).tolist())
            for text in batch:
                n_plain = len(tokenizer(text, add_special_tokens=False, truncation=False)["input_ids"])
                if n_plain + tokenizer.num_special_tokens_to_add(pair=False) > args.max_length:
                    truncated_rows += 1

            encoded_dev = {k: v.to(device) for k, v in encoded.items()}
            outputs = model(**encoded_dev, return_dict=True)
            hidden = outputs.last_hidden_state
            pooling = spec["pooling"]
            if pooling == "mean_non_special":
                mask = attention.bool() & ~special_mask.bool()
                pooled = masked_mean(hidden, mask.to(device))
            elif pooling == "mean_attention":
                pooled = masked_mean(hidden, attention.bool().to(device))
            elif pooling == "cls":
                pooled = hidden[:, 0, :]
            elif pooling == "last_token":
                # Works for right- or left-padded batches by selecting the final attended token.
                att = attention.to(device)
                seq = torch.arange(att.shape[1], device=device).unsqueeze(0).expand_as(att)
                last_idx = (seq * att).argmax(dim=1)
                pooled = hidden[torch.arange(hidden.shape[0], device=device), last_idx]
            else:
                raise RuntimeError(f"Unknown pooling: {pooling}")

            if spec["normalize"]:
                pooled = l2_normalize(pooled)
            chunks.append(pooled.cpu().numpy().astype(np.float32))

    embeddings = np.concatenate(chunks, axis=0)
    if embeddings.shape[0] != len(texts) or not np.isfinite(embeddings).all():
        raise RuntimeError("Invalid embedding output")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = (args.output_dir / args.model_key / f"run-{args.run_number:02d}" / stamp).resolve()
    out.mkdir(parents=True, exist_ok=False)
    np.save(out / "embeddings.npy", embeddings)
    (out / "texts.json").write_text(json.dumps(texts, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "dataset": str(dataset),
        "workbook": str(workbook),
        "run_number": args.run_number,
        "model_key": args.model_key,
        "model_id": model_id,
        "model_revision": revision,
        "transformers_version": transformers.__version__,
        "torch_version": torch.__version__,
        "device": device,
        "batch_size": args.batch_size,
        "max_length": args.max_length,
        "pooling": spec["pooling"],
        "input_prefix": spec["prefix"],
        "l2_normalized": spec["normalize"],
        "n_rows": len(texts),
        "embedding_shape": list(embeddings.shape),
        "min_attention_tokens": int(min(token_counts)),
        "max_attention_tokens": int(max(token_counts)),
        "truncated_rows": int(truncated_rows),
        "screening_policy": "runs 01-06 only; run 07 sealed until model selection is frozen",
    }
    (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Model-panel embedding output: {out}")
    print(f"Model: {model_id}@{revision}")
    print(f"Run: {args.run_number:02d} | rows: {len(texts)} | shape: {embeddings.shape} | device: {device}")
    print(f"Pooling: {spec['pooling']} | prefix: {spec['prefix']!r} | truncated rows: {truncated_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
