#!/usr/bin/env python3
"""Generate frozen NeuroSem embeddings for the sealed ChineseEEG run-07 holdout.

This script is intentionally restricted to the three model representations selected
before run-07 semantic results were accessed. Exact model revisions and pooling rules
are hard-coded from the completed runs 01-06 screen.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

FROZEN_SPECS = {
    "bert-base-chinese-final": {
        "model_id": "google-bert/bert-base-chinese",
        "revision": "8d2a91f91cc38c96bb8b4556ba70c392f8d5ee55",
        "pooling": "mean_non_special",
        "prefix": "",
        "normalize": False,
    },
    "multilingual-e5-large": {
        "model_id": "intfloat/multilingual-e5-large",
        "revision": "3d7cfbdacd47fdda877c5cd8a79fbcc4f2a574f3",
        "pooling": "mean_attention",
        "prefix": "query: ",
        "normalize": True,
    },
    "bge-m3": {
        "model_id": "BAAI/bge-m3",
        "revision": "5617a9f61b028005a4858fdac845db406aefb181",
        "pooling": "cls",
        "prefix": "",
        "normalize": True,
    },
}


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[str] = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    denom = mask.sum(dim=1).clamp_min(1.0)
    return (hidden * mask).sum(dim=1) / denom


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate frozen embeddings for ChineseEEG run-07 holdout.")
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--model-key", choices=sorted(FROZEN_SPECS), required=True)
    parser.add_argument("--batch-size", type=int, default=128)
    parser.add_argument("--max-length", type=int, default=64)
    parser.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/chineseeeg_run07_holdout_embeddings"))
    args = parser.parse_args()

    if args.batch_size < 1:
        raise SystemExit("--batch-size must be >= 1")

    try:
        import torch
        import transformers
        from transformers import AutoModel, AutoTokenizer
    except ImportError as exc:
        raise SystemExit("Missing embedding dependencies. Install requirements-embeddings.txt") from exc

    spec = FROZEN_SPECS[args.model_key]
    dataset = args.dataset.expanduser().resolve()
    workbook = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / "segmented_Chinense_novel_run_7.xlsx"
    if not workbook.exists():
        raise SystemExit(f"Run-07 workbook not materialized: {workbook}")

    texts = read_rows(workbook)
    if not texts:
        raise SystemExit("No run-07 workbook rows found")
    model_texts = [spec["prefix"] + t for t in texts]

    if args.device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    else:
        device = args.device
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("CUDA requested but unavailable")

    tokenizer = AutoTokenizer.from_pretrained(
        spec["model_id"], revision=spec["revision"], trust_remote_code=True
    )
    model = AutoModel.from_pretrained(
        spec["model_id"], revision=spec["revision"], trust_remote_code=True
    )
    model.eval().to(device)

    chunks = []
    token_counts: list[int] = []
    truncated_rows = 0

    with torch.inference_mode():
        for start in range(0, len(model_texts), args.batch_size):
            batch = model_texts[start:start + args.batch_size]
            encoded = tokenizer(
                batch,
                padding=True,
                truncation=True,
                max_length=args.max_length,
                return_tensors="pt",
                return_special_tokens_mask=True,
            )
            special_mask = encoded.pop("special_tokens_mask")
            attention = encoded["attention_mask"]
            token_counts.extend(attention.sum(dim=1).tolist())
            for text in batch:
                n_plain = len(tokenizer(text, add_special_tokens=False, truncation=False)["input_ids"])
                if n_plain + tokenizer.num_special_tokens_to_add(pair=False) > args.max_length:
                    truncated_rows += 1

            encoded_dev = {k: v.to(device) for k, v in encoded.items()}
            outputs = model(**encoded_dev, return_dict=True)
            hidden = outputs.last_hidden_state

            if spec["pooling"] == "mean_non_special":
                mask = attention.bool() & ~special_mask.bool()
                pooled = masked_mean(hidden, mask.to(device))
            elif spec["pooling"] == "mean_attention":
                pooled = masked_mean(hidden, attention.bool().to(device))
            elif spec["pooling"] == "cls":
                pooled = hidden[:, 0, :]
            else:
                raise RuntimeError(f"Unsupported frozen pooling: {spec['pooling']}")

            if spec["normalize"]:
                pooled = torch.nn.functional.normalize(pooled, p=2, dim=1)
            chunks.append(pooled.cpu().numpy().astype(np.float32))

    embeddings = np.concatenate(chunks, axis=0)
    if embeddings.shape[0] != len(texts) or not np.isfinite(embeddings).all():
        raise RuntimeError("Invalid run-07 embedding output")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = (args.output_dir / args.model_key / "run-07" / stamp).resolve()
    out.mkdir(parents=True, exist_ok=False)
    np.save(out / "embeddings.npy", embeddings)
    (out / "texts.json").write_text(json.dumps(texts, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "dataset": str(dataset),
        "workbook": str(workbook),
        "run_number": 7,
        "model_key": args.model_key,
        "model_id": spec["model_id"],
        "model_revision": spec["revision"],
        "transformers_version": transformers.__version__,
        "torch_version": torch.__version__,
        "device": device,
        "batch_size": args.batch_size,
        "max_length": args.max_length,
        "pooling": spec["pooling"],
        "input_prefix": spec["prefix"],
        "l2_normalized": spec["normalize"],
        "n_rows": len(texts),
        "embedding_shape": list(embeddings.shape),
        "min_attention_tokens": int(min(token_counts)),
        "max_attention_tokens": int(max(token_counts)),
        "truncated_rows": int(truncated_rows),
        "holdout_policy": "run-07 opened only after model selection was frozen and logged",
    }
    (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Run-07 holdout embedding output: {out}")
    print(f"Model: {spec['model_id']}@{spec['revision']}")
    print(f"Run: 07 | rows: {len(texts)} | shape: {embeddings.shape} | device: {device}")
    print(f"Pooling: {spec['pooling']} | prefix: {spec['prefix']!r} | truncated rows: {truncated_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
