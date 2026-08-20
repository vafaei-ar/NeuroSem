#!/usr/bin/env python3
"""Generate reproducible ChineseEEG LittlePrince row embeddings from a pinned BERT revision.

Primary target: mean-pooled final hidden state over non-special, non-padding tokens.
Sensitivity target: mean of the last four hidden layers, then mean-pooled over the same tokens.

The script reads the canonical segmented workbook, preserves all 395 rows including the
four numeric chapter rows, and writes embeddings plus provenance metadata. Semantic
analyses should select the 391 semantic-eligible rows by the existing manifest indices.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

MODEL_ID = "google-bert/bert-base-chinese"
MODEL_REVISION = "8d2a91f91cc38c96bb8b4556ba70c392f8d5ee55"


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[str] = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    import torch

    mask = mask.to(hidden.dtype).unsqueeze(-1)
    denom = mask.sum(dim=1).clamp_min(1.0)
    return (hidden * mask).sum(dim=1) / denom


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate pinned BERT embeddings for ChineseEEG LittlePrince rows.")
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--run-number", type=int, default=1)
    parser.add_argument("--model", default=MODEL_ID)
    parser.add_argument("--revision", default=MODEL_REVISION)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--max-length", type=int, default=64)
    parser.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/chineseeeg_pinned_embeddings"))
    args = parser.parse_args()

    if args.batch_size < 1:
        raise SystemExit("--batch-size must be >= 1")
    if args.max_length < 8:
        raise SystemExit("--max-length must be >= 8")

    try:
        import torch
        import transformers
        from transformers import AutoModel, AutoTokenizer
    except ImportError as exc:
        raise SystemExit(
            "Missing embedding dependencies. Install with: pip install -r requirements-embeddings.txt"
        ) from exc

    dataset = args.dataset.expanduser().resolve()
    workbook = (
        dataset
        / "derivatives"
        / "novels"
        / "segmented_novel"
        / "LittlePrince"
        / f"segmented_Chinense_novel_run_{args.run_number}.xlsx"
    )
    if not workbook.exists():
        raise SystemExit(f"Workbook not materialized: {workbook}")

    texts = read_rows(workbook)
    if not texts:
        raise SystemExit("No non-empty workbook rows found")

    if args.device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    else:
        device = args.device
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("--device cuda requested but CUDA is unavailable")

    tokenizer = AutoTokenizer.from_pretrained(args.model, revision=args.revision)
    model = AutoModel.from_pretrained(args.model, revision=args.revision)
    model.eval().to(device)

    final_chunks = []
    last4_chunks = []
    token_counts: list[int] = []
    truncated_rows = 0

    special_ids = set(tokenizer.all_special_ids)

    with torch.inference_mode():
        for start in range(0, len(texts), args.batch_size):
            batch = texts[start:start + args.batch_size]
            encoded = tokenizer(
                batch,
                padding=True,
                truncation=True,
                max_length=args.max_length,
                return_tensors="pt",
                return_special_tokens_mask=True,
            )
            special_mask = encoded.pop("special_tokens_mask")
            attention = encoded["attention_mask"]
            token_mask = attention.bool() & ~special_mask.bool()
            token_counts.extend(token_mask.sum(dim=1).tolist())

            # Detect likely truncation by separately tokenizing without special tokens.
            for text in batch:
                n_plain = len(tokenizer(text, add_special_tokens=False, truncation=False)["input_ids"])
                if n_plain + tokenizer.num_special_tokens_to_add(pair=False) > args.max_length:
                    truncated_rows += 1

            encoded = {k: v.to(device) for k, v in encoded.items()}
            token_mask_dev = token_mask.to(device)
            outputs = model(**encoded, output_hidden_states=True, return_dict=True)

            final = masked_mean(outputs.last_hidden_state, token_mask_dev)
            hidden_states = outputs.hidden_states
            if hidden_states is None or len(hidden_states) < 4:
                raise RuntimeError("Model did not return enough hidden layers for last-four sensitivity embedding")
            last4 = torch.stack(hidden_states[-4:], dim=0).mean(dim=0)
            last4 = masked_mean(last4, token_mask_dev)

            final_chunks.append(final.cpu().numpy().astype(np.float32))
            last4_chunks.append(last4.cpu().numpy().astype(np.float32))

    final_embeddings = np.concatenate(final_chunks, axis=0)
    last4_embeddings = np.concatenate(last4_chunks, axis=0)

    if final_embeddings.shape[0] != len(texts) or last4_embeddings.shape[0] != len(texts):
        raise RuntimeError("Embedding row count mismatch")
    if not (np.isfinite(final_embeddings).all() and np.isfinite(last4_embeddings).all()):
        raise RuntimeError("Non-finite generated embeddings")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = (args.output_dir / stamp).resolve()
    out.mkdir(parents=True, exist_ok=False)

    np.save(out / "bert_base_chinese_final_mean.npy", final_embeddings)
    np.save(out / "bert_base_chinese_last4_mean.npy", last4_embeddings)
    (out / "texts.json").write_text(json.dumps(texts, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "dataset": str(dataset),
        "workbook": str(workbook),
        "run_number": args.run_number,
        "n_rows": len(texts),
        "model_id": args.model,
        "model_revision": args.revision,
        "transformers_version": transformers.__version__,
        "torch_version": torch.__version__,
        "device": device,
        "batch_size": args.batch_size,
        "max_length": args.max_length,
        "token_pooling": "mean over non-special, non-padding tokens",
        "primary_embedding": "final hidden layer mean pooling",
        "sensitivity_embedding": "mean of final four hidden layers, then mean pooling",
        "final_shape": list(final_embeddings.shape),
        "last4_shape": list(last4_embeddings.shape),
        "min_non_special_tokens": int(min(token_counts)),
        "max_non_special_tokens": int(max(token_counts)),
        "truncated_rows": int(truncated_rows),
        "notes": [
            "All workbook rows are preserved, including numeric chapter rows, to keep canonical indices.",
            "The primary semantic analysis must select semantic-eligible rows using embedding_index from the verified manifests.",
            "This independently generated representation is intentionally different from the provenance-uncertain distributed author embeddings.",
        ],
    }
    (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Pinned embedding output: {out}")
    print(f"Model: {args.model}@{args.revision}")
    print(f"Rows: {len(texts)} | device: {device} | transformers: {transformers.__version__}")
    print(f"Final-layer mean shape: {final_embeddings.shape}")
    print(f"Last-four mean shape: {last4_embeddings.shape}")
    print(f"Non-special tokens: min={min(token_counts)} max={max(token_counts)}")
    print(f"Rows truncated at max_length={args.max_length}: {truncated_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
