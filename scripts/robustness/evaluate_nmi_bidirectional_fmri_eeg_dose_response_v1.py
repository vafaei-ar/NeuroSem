#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import os
import string
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl
from scipy.spatial.distance import pdist
from scipy.stats import rankdata, spearmanr

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.analysis.run_zuco2_nr_primary_representation_reliability import (
    EXPECTED,
    fisher_mean,
    load_inventory,
    load_material_rows,
    load_run_features,
    nuisance_matrix,
    rdm_edges,
    residualize,
)
from scripts.tuning.evaluate_tmnred_e5_transfer_v1 import encode_texts, load_adapter

CAL_ROOT = Path("outputs/nmi_bidirectional_fmri_calibration_v1/latest")
LAMBDAS = [0.0, 0.01, 0.03, 0.10, 0.30, 1.0]
POSITIVE = [0.01, 0.03, 0.10, 0.30, 1.0]
LAMBDA_DIR = {0.0: "lambda_0p0", 0.01: "lambda_0p01", 0.03: "lambda_0p03", 0.10: "lambda_0p1", 0.30: "lambda_0p3", 1.0: "lambda_1p0"}
MODEL_ID = "intfloat/multilingual-e5-large"
MODEL_REVISION = "3d7cfbdacd47fdda877c5cd8a79fbcc4f2a574f3"
PREFIX = "query: "
CHINESE_SUBJECTS = ["sub-04", "sub-05", "sub-06", "sub-07", "sub-08", "sub-09", "sub-10", "sub-13", "sub-14", "sub-15"]
BOOT_SEED = 20260830
N_BOOT = 10000


def report_progress(current: int, total: int, phase: str) -> None:
    raw = os.environ.get("RUNRELAY_PROGRESS_FILE")
    if not raw:
        return
    p = Path(raw); p.parent.mkdir(parents=True, exist_ok=True)
    d = {"schema_version": 1, "current": current, "total": total, "fraction": current / total, "phase": phase, "unit": "model-target arms", "updated_at_epoch": time.time()}
    t = p.with_suffix(p.suffix + ".tmp"); t.write_text(json.dumps(d), encoding="utf-8"); os.replace(t, p)


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)


def adapter_for(lam: float) -> Path:
    p = CAL_ROOT / LAMBDA_DIR[lam] / "adapter"
    if not p.exists():
        raise FileNotFoundError(p)
    return p


def assert_source_freeze() -> dict:
    p = CAL_ROOT / "summary.json"
    if not p.exists(): raise FileNotFoundError(p)
    s = json.loads(p.read_text(encoding="utf-8"))
    if s.get("source_gate_pass") is not True: raise RuntimeError("source gate did not pass")
    if abs(float(s.get("selected_lambda")) - 0.01) > 1e-12: raise RuntimeError("unexpected frozen primary lambda")
    for lam in LAMBDAS: adapter_for(lam)
    return s


def exact_signflip(diffs: np.ndarray) -> dict:
    diffs = np.asarray(diffs, float); n = len(diffs); obs = float(diffs.mean())
    vals = np.empty(1 << n, float)
    for mask in range(1 << n):
        signs = np.array([1.0 if (mask >> i) & 1 else -1.0 for i in range(n)])
        vals[mask] = float(np.mean(diffs * signs))
    return {"observed_mean": obs, "one_sided_greater_p": float(np.mean(vals >= obs - 1e-15)), "two_sided_p": float(np.mean(np.abs(vals) >= abs(obs) - 1e-15)), "n_sign_patterns": int(1 << n)}


def bootstrap_ci(x: np.ndarray, seed_offset: int = 0) -> list[float]:
    x = np.asarray(x, float); rng = np.random.default_rng(BOOT_SEED + seed_offset)
    vals = np.empty(N_BOOT, float)
    for i in range(N_BOOT): vals[i] = rng.choice(x, size=len(x), replace=True).mean()
    return [float(np.percentile(vals, 2.5)), float(np.percentile(vals, 97.5))]


def summarize_diffs(diffs: np.ndarray, seed_offset: int = 0) -> dict:
    diffs = np.asarray(diffs, float)
    return {"mean_delta": float(diffs.mean()), "median_delta": float(np.median(diffs)), "n_positive": int(np.sum(diffs > 0)), "fraction_positive": float(np.mean(diffs > 0)), "bootstrap_95ci": bootstrap_ci(diffs, seed_offset), "exact_signflip": exact_signflip(diffs)}


def holm_adjust(ps: list[float]) -> list[float]:
    m = len(ps); order = np.argsort(ps); out = np.empty(m, float); running = 0.0
    for rank, idx in enumerate(order):
        adj = min(1.0, (m-rank) * ps[idx]); running = max(running, adj); out[idx] = running
    return [float(x) for x in out]


def trend_summary(by_lambda: dict[float, np.ndarray], seed_offset: int = 50) -> dict:
    x = np.log10(np.asarray(POSITIVE, float)); X = np.column_stack([np.ones(len(x)), x]); pinv = np.linalg.pinv(X)
    mat = np.column_stack([by_lambda[lam] for lam in POSITIVE])
    slopes = np.asarray([(pinv @ row)[1] for row in mat], float)
    target_means = np.asarray([by_lambda[lam].mean() for lam in POSITIVE], float)
    return {"x": "log10(lambda)", "participant_slope_summary": summarize_diffs(slopes, seed_offset), "participant_slopes": slopes.tolist(), "descriptive_spearman_mean_delta_vs_log10_lambda": float(spearmanr(x, target_means).statistic), "target_mean_deltas": {str(lam): float(by_lambda[lam].mean()) for lam in POSITIVE}}


def safe_rho(a, b) -> float:
    r = float(spearmanr(a, b).statistic)
    if not np.isfinite(r): raise RuntimeError("non-finite Spearman RSA")
    return r


def evaluate_zuco(device: str, out: Path, progress_start: int) -> dict:
    data_root = Path("data/raw/zuco2_nr"); input_freeze = Path("outputs/zuco2_nr_input_materialization/latest/summary.json"); mapping_freeze = Path("outputs/zuco2_nr_format_probe/latest/summary.json"); stimulus_root = Path("data/raw/zuco2_probe")
    cohort = json.loads(input_freeze.read_text()); mapping = json.loads(mapping_freeze.read_text())
    ready = list(cohort.get("ready_subjects_all_7_runs") or [])
    if len(ready) != 17 or cohort.get("n_ready_subjects_all_7_runs") != 17 or "YTL" in ready: raise RuntimeError("unexpected ZuCo cohort")
    maps = {r["run"]: r for r in mapping.get("wordcount_mapping_diagnostics", [])}
    inventory = load_inventory(input_freeze.parent / "session_inventory.csv"); path_by = {}
    for r in inventory:
        if r.get("subject") in ready and str(r.get("ready", "")).lower() == "true": path_by[(r["subject"], int(r["run"]))] = data_root.resolve() / r["osf_path"]
    texts_by_run = {}; nuis_by_run = {}
    for run in range(1, 8):
        rows = load_material_rows(stimulus_root / "task_materials" / f"nr_{run}.csv"); selected = maps[f"NR{run}"]["selected_material_rows_1based"]
        texts = [str(rows[i-1][2]).strip() for i in selected]; texts_by_run[run] = texts; nuis_by_run[run] = nuisance_matrix(texts)
    flat = [t for run in range(1, 8) for t in texts_by_run[run]]
    model_edges = {}; provenance = {}
    import torch
    for j, lam in enumerate(LAMBDAS):
        tok, model = load_adapter(adapter_for(lam), device); emb = encode_texts(model, tok, flat, device); model_edges[lam] = {}; off = 0
        for run in range(1, 8):
            n = EXPECTED[run]; model_edges[lam][run] = pdist(emb[off:off+n], metric="cosine"); off += n
        provenance[str(lam)] = str(adapter_for(lam).resolve()); del model; torch.cuda.empty_cache(); report_progress(progress_start+j+1, 12, f"ZuCo model arm {j+1}/6")
    neural_resid = {}
    for sub in ready:
        neural_resid[sub] = {}
        for run in range(1, 8):
            feats, _ = load_run_features(path_by[(sub, run)], EXPECTED[run]); neural_resid[sub][run] = residualize(rdm_edges(feats["row_mean_all"]), nuis_by_run[run])
    arm_subject = {lam: {} for lam in LAMBDAS}; rows = []
    for sub in ready:
        for lam in LAMBDAS:
            vals = [safe_rho(neural_resid[sub][run], residualize(model_edges[lam][run], nuis_by_run[run])) for run in range(1, 8)]
            arm_subject[lam][sub] = fisher_mean(vals)
    baseline = np.asarray([arm_subject[0.0][s] for s in ready], float); by_lambda = {}; contrasts = []
    for i, lam in enumerate(POSITIVE):
        vals = np.asarray([arm_subject[lam][s] for s in ready], float); d = vals-baseline; by_lambda[lam] = d; rec = {"lambda": lam, **summarize_diffs(d, i)}; contrasts.append(rec)
        for k, sub in enumerate(ready): rows.append({"target":"ZuCo","subject":sub,"lambda":lam,"lambda_0_rsa":baseline[k],"lambda_rsa":vals[k],"delta":d[k]})
    adj = holm_adjust([r["exact_signflip"]["one_sided_greater_p"] for r in contrasts])
    for r, a in zip(contrasts, adj): r["holm_adjusted_one_sided_p_within_target"] = a
    write_csv(out / "zuco_subject_dose_results.csv", rows)
    return {"n_subjects": len(ready), "contrasts": contrasts, "trend": trend_summary(by_lambda, 100), "model_provenance": provenance}


def read_chinese_texts(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active; out=[]
    for r in range(2, ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v is not None: out.append(str(v))
    return out


def rank_z(x: np.ndarray) -> np.ndarray:
    r=rankdata(np.asarray(x,float),method="average"); r-=r.mean(); sd=r.std(); return r/sd


def residualize_ranked(y: np.ndarray, nuisances: list[np.ndarray]) -> np.ndarray:
    yr=rank_z(y); X=np.column_stack([np.ones_like(yr), *[rank_z(n) for n in nuisances]]); beta,*_=np.linalg.lstsq(X,yr,rcond=None); resid=yr-X@beta; resid-=resid.mean(); return resid/resid.std()


def char_set_jaccard_rdm(texts: list[str]) -> np.ndarray:
    sets=[set(t) for t in texts]; out=[]
    for i in range(len(sets)-1):
        for j in range(i+1,len(sets)):
            u=len(sets[i]|sets[j]); out.append(1.0-(len(sets[i]&sets[j])/u if u else 1.0))
    return np.asarray(out,float)


def punctuation_count(text: str) -> int:
    punct=set(string.punctuation)|set("，。！？；：、“”‘’（）《》〈〉【】…—·"); return sum(ch in punct for ch in text)


def evaluate_chinese(device: str, out: Path, progress_start: int) -> dict:
    import torch
    from peft import PeftModel
    from transformers import AutoModel, AutoTokenizer
    from scripts.analysis.assess_chineseeeg_run07_holdout_fast import read_csv, latest_feature_dir, zscore_columns

    data_root=Path("data/raw/chineseeeg").resolve(); workbook=data_root/"derivatives"/"novels"/"segmented_novel"/"LittlePrince"/"segmented_Chinense_novel_run_7.xlsx"; texts_all=read_chinese_texts(workbook)
    tok=AutoTokenizer.from_pretrained(MODEL_ID,revision=MODEL_REVISION); embeddings={}; provenance={}
    for j,lam in enumerate(LAMBDAS):
        base=AutoModel.from_pretrained(MODEL_ID,revision=MODEL_REVISION); model=PeftModel.from_pretrained(base,adapter_for(lam)).to(device)
        chunks=[]; model.eval()
        with torch.inference_mode():
            for start in range(0,len(texts_all),64):
                b=[PREFIX+t for t in texts_all[start:start+64]]; e=tok(b,padding=True,truncation=True,max_length=64,return_tensors="pt"); att=e["attention_mask"].to(device); e={k:v.to(device) for k,v in e.items()}; o=model(**e,return_dict=True); m=att.to(o.last_hidden_state.dtype).unsqueeze(-1); x=(o.last_hidden_state*m).sum(1)/m.sum(1).clamp_min(1.); x=torch.nn.functional.normalize(x,p=2,dim=1); chunks.append(x.cpu().numpy())
        embeddings[lam]=np.concatenate(chunks,0); provenance[str(lam)]=str(adapter_for(lam).resolve()); del model,base; torch.cuda.empty_cache(); report_progress(progress_start+j+1,12,f"ChineseEEG model arm {j+1}/6")
    ref_texts=None; ref_idx=None; subject_rsa={lam:{} for lam in LAMBDAS}
    for sub in CHINESE_SUBJECTS:
        d=latest_feature_dir(Path("outputs/chineseeeg_row_features/run-07"),sub); meta=read_csv(d/"metadata.csv"); x=np.load(d/"row_mean.npy").astype(float); texts=[r["text"] for r in meta]; idx=np.asarray([int(r["embedding_index"]) for r in meta],int); chapters=np.asarray([int((r["chapter_marker_context"] or "CH00")[2:]) for r in meta],int)
        if ref_texts is None: ref_texts=texts; ref_idx=idx
        elif texts!=ref_texts or not np.array_equal(idx,ref_idx): raise RuntimeError("ChineseEEG canonical row mismatch")
        orth=char_set_jaccard_rdm(texts); punct=np.asarray([punctuation_count(t) for t in texts],float); nuis=[pdist(np.asarray([float(r["run_position_fraction"]) for r in meta])[:,None],metric="cityblock"),pdist(np.asarray([float(r["duration_sec"]) for r in meta])[:,None],metric="cityblock"),pdist(np.asarray([float(r["char_count"]) for r in meta])[:,None],metric="cityblock"),pdist(chapters[:,None],metric="hamming"),orth,pdist(punct[:,None],metric="cityblock")]
        nr=residualize_ranked(pdist(zscore_columns(x),metric="correlation"),nuis)
        for lam in LAMBDAS:
            sr=residualize_ranked(pdist(embeddings[lam][idx],metric="cosine"),nuis); subject_rsa[lam][sub]=float(np.mean(nr*sr))
    baseline=np.asarray([subject_rsa[0.0][s] for s in CHINESE_SUBJECTS],float); by_lambda={}; contrasts=[]; rows=[]
    for i,lam in enumerate(POSITIVE):
        vals=np.asarray([subject_rsa[lam][s] for s in CHINESE_SUBJECTS],float); d=vals-baseline; by_lambda[lam]=d; rec={"lambda":lam,**summarize_diffs(d,200+i)}; contrasts.append(rec)
        for k,sub in enumerate(CHINESE_SUBJECTS): rows.append({"target":"ChineseEEG_run07","subject":sub,"lambda":lam,"lambda_0_rsa":baseline[k],"lambda_rsa":vals[k],"delta":d[k]})
    adj=holm_adjust([r["exact_signflip"]["one_sided_greater_p"] for r in contrasts])
    for r,a in zip(contrasts,adj): r["holm_adjusted_one_sided_p_within_target"]=a
    write_csv(out/"chineseeeg_subject_dose_results.csv",rows)
    return {"n_subjects":len(CHINESE_SUBJECTS),"contrasts":contrasts,"trend":trend_summary(by_lambda,300),"model_provenance":provenance}


def main() -> int:
    import torch
    source=assert_source_freeze(); device="cuda" if torch.cuda.is_available() else "cpu"
    if device!="cuda": raise RuntimeError("GPU required")
    out=Path("outputs/nmi_bidirectional_fmri_eeg_dose_response_v1/latest").resolve(); out.mkdir(parents=True,exist_ok=True)
    zuco=evaluate_zuco(device,out,0); chinese=evaluate_chinese(device,out,6)
    payload={"schema_version":1,"analysis_stage":"exploratory/post-confirmatory fMRI-to-EEG dose-response characterization","protocol":"docs/21_NMI_BIDIRECTIONAL_FMRI_EEG_DOSE_RESPONSE_V1.md","source_dataset":"SMN4Lang fMRI","model_id":MODEL_ID,"model_revision":MODEL_REVISION,"lambda_grid":LAMBDAS,"frozen_primary_reverse_transfer_lambda":0.01,"source_selection_rule":source.get("selection_rule"),"zuco":zuco,"chineseeeg_run07":chinese,"guardrails":["No models were retrained and no new lambda values were added.","The lambda=.01 ZuCo result remains the primary frozen reverse-transfer test.","All dose-response results are exploratory/post-confirmatory and cannot be used to redefine a confirmatory optimum.","Every already-trained positive lambda is reported on both EEG targets."]}
    (out/"summary.json").write_text(json.dumps(payload,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"zuco_contrasts":zuco["contrasts"],"zuco_trend":zuco["trend"],"chineseeeg_contrasts":chinese["contrasts"],"chineseeeg_trend":chinese["trend"]},indent=2))
    return 0

if __name__=="__main__": raise SystemExit(main())
