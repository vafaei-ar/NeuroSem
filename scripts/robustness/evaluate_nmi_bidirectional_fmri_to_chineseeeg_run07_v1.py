#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

MODEL_ID = "intfloat/multilingual-e5-large"
MODEL_REVISION = "3d7cfbdacd47fdda877c5cd8a79fbcc4f2a574f3"
PREFIX = "query: "
SUBJECTS = ["sub-04", "sub-05", "sub-06", "sub-07", "sub-08", "sub-09", "sub-10", "sub-13", "sub-14", "sub-15"]
BOOT_SEED = 20260830
N_BOOT = 10000


def report_progress(current: int, total: int, phase: str) -> None:
    raw = os.environ.get("RUNRELAY_PROGRESS_FILE")
    if not raw:
        return
    p = Path(raw)
    p.parent.mkdir(parents=True, exist_ok=True)
    d = {"schema_version": 1, "current": current, "total": total, "fraction": current / total, "phase": phase, "unit": "model-arms", "updated_at_epoch": time.time()}
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(d), encoding="utf-8")
    os.replace(tmp, p)


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    return (hidden * mask).sum(dim=1) / mask.sum(dim=1).clamp_min(1.0)


def generate_embeddings(model, tokenizer, texts, device, batch_size=64, max_length=64):
    import torch
    chunks = []
    model.eval()
    with torch.inference_mode():
        for start in range(0, len(texts), batch_size):
            batch = [PREFIX + t for t in texts[start:start + batch_size]]
            enc = tokenizer(batch, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
            attention = enc["attention_mask"]
            enc = {k: v.to(device) for k, v in enc.items()}
            out = model(**enc, return_dict=True)
            pooled = masked_mean(out.last_hidden_state, attention.to(device).bool())
            pooled = torch.nn.functional.normalize(pooled, p=2, dim=1)
            chunks.append(pooled.cpu().numpy().astype(np.float32))
    return np.concatenate(chunks, axis=0)


def latest_summary(root: Path) -> Path:
    xs = [d for d in root.iterdir() if d.is_dir() and (d / "summary.json").exists()] if root.exists() else []
    if not xs:
        raise FileNotFoundError(f"No summary under {root}")
    return sorted(xs)[-1] / "summary.json"


def exact_signflip(diffs: np.ndarray) -> dict:
    n = len(diffs)
    obs = float(diffs.mean())
    vals = []
    for mask in range(1 << n):
        signs = np.array([1.0 if (mask >> i) & 1 else -1.0 for i in range(n)])
        vals.append(float(np.mean(diffs * signs)))
    vals = np.asarray(vals)
    return {
        "observed_mean": obs,
        "one_sided_greater_p": float(np.mean(vals >= obs - 1e-15)),
        "two_sided_p": float(np.mean(np.abs(vals) >= abs(obs) - 1e-15)),
        "n_sign_patterns": int(1 << n),
    }


def bootstrap_ci(diffs: np.ndarray) -> list[float]:
    rng = np.random.default_rng(BOOT_SEED)
    means = np.empty(N_BOOT, float)
    for i in range(N_BOOT):
        means[i] = rng.choice(diffs, size=len(diffs), replace=True).mean()
    return [float(np.percentile(means, 2.5)), float(np.percentile(means, 97.5))]


def main() -> int:
    import torch
    from peft import PeftModel
    from transformers import AutoModel, AutoTokenizer

    data_root = Path("data/raw/chineseeeg").resolve()
    workbook = data_root / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / "segmented_Chinense_novel_run_7.xlsx"
    if not workbook.exists():
        raise FileNotFoundError(workbook)
    texts = read_rows(workbook)
    if not texts:
        raise RuntimeError("No run-07 text rows")

    device = "cuda" if torch.cuda.is_available() else "cpu"
    if device != "cuda":
        raise RuntimeError("GPU required")

    calibration = Path("outputs/nmi_bidirectional_fmri_calibration_v1/latest").resolve()
    arms = {
        "lambda_0": calibration / "lambda_0p0" / "adapter",
        "fmri_guided_lambda_0p01": calibration / "lambda_0p01" / "adapter",
    }
    for p in arms.values():
        if not p.exists():
            raise FileNotFoundError(p)

    out = Path("outputs/nmi_bidirectional_fmri_to_chineseeeg_run07_v1/latest").resolve()
    emb_root = out / "embeddings"
    rsa_root = out / "rsa"
    out.mkdir(parents=True, exist_ok=True)
    tok = AutoTokenizer.from_pretrained(MODEL_ID, revision=MODEL_REVISION)

    summaries = {}
    for ai, (arm, adapter) in enumerate(arms.items(), 1):
        base = AutoModel.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
        model = PeftModel.from_pretrained(base, adapter).to(device)
        emb = generate_embeddings(model, tok, texts, device)
        arm_emb = emb_root / arm / datetime.now().strftime("%Y%m%d_%H%M%S")
        arm_emb.mkdir(parents=True, exist_ok=False)
        np.save(arm_emb / "embeddings.npy", emb)
        (arm_emb / "summary.json").write_text(json.dumps({
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "arm": arm,
            "model_id": MODEL_ID,
            "model_revision": MODEL_REVISION,
            "adapter_dir": str(adapter),
            "run_number": 7,
            "n_rows": len(texts),
            "embedding_shape": list(emb.shape),
            "pooling": "attention-mask mean, L2 normalized",
            "input_prefix": PREFIX,
            "max_length": 64,
            "evaluation_only": True,
        }, indent=2), encoding="utf-8")

        cmd = [
            sys.executable,
            "scripts/analysis/assess_chineseeeg_run07_holdout_fast.py",
            "--embedding-root", str(emb_root / arm),
            "--feature-root", "outputs/chineseeeg_row_features/run-07",
            "--subjects", *SUBJECTS,
            "--permutations", "10000",
            "--workers", "16",
            "--chunk-size", "50",
            "--output-dir", str(rsa_root / arm),
        ]
        subprocess.run(cmd, check=True)
        sp = latest_summary(rsa_root / arm)
        summaries[arm] = json.loads(sp.read_text(encoding="utf-8"))
        del model, base
        torch.cuda.empty_cache()
        report_progress(ai, 2, f"ChineseEEG run-07 arm {ai}/2")

    by0 = summaries["lambda_0"]["observed"]["by_subject"]
    by1 = summaries["fmri_guided_lambda_0p01"]["observed"]["by_subject"]
    rows = []
    diffs = []
    for sub in SUBJECTS:
        a0 = float(by0[sub]); a1 = float(by1[sub]); d = a1 - a0
        diffs.append(d)
        rows.append({"subject": sub, "lambda_0_resid_rsa": a0, "fmri_guided_lambda_0p01_resid_rsa": a1, "delta_fmri_guided_minus_0": d})
    diffs = np.asarray(diffs, float)

    import csv
    with (out / "subject_results.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)

    payload = {
        "schema_version": 1,
        "analysis_stage": "post-confirmatory bidirectional cross-modal transfer secondary consistency target",
        "protocol": "docs/20_NMI_BIDIRECTIONAL_FMRI_TO_CHINESEEEG_RUN07_V1.md",
        "source_dataset": "SMN4Lang fMRI",
        "target_dataset": "ChineseEEG sealed run-07 EEG",
        "selected_lambda": 0.01,
        "n_frozen_subjects": len(SUBJECTS),
        "primary_secondary_result": {
            "contrast": "fmri_guided_lambda_0p01_minus_lambda_0",
            "mean_delta": float(diffs.mean()),
            "median_delta": float(np.median(diffs)),
            "fraction_subjects_positive": float(np.mean(diffs > 0)),
            "n_subjects_positive": int(np.sum(diffs > 0)),
            "bootstrap_95ci": bootstrap_ci(diffs),
            "bootstrap_seed": BOOT_SEED,
            "n_bootstrap": N_BOOT,
            "exact_signflip": exact_signflip(diffs),
        },
        "arm_run07_summaries": {
            arm: {"mean": float(s["observed"]["mean"]), "median": float(s["observed"]["median"]), "within_chapter_permutation_p": float(s["inference"]["p_ge_observed"])}
            for arm, s in summaries.items()
        },
        "guardrails": [
            "The fMRI-guided lambda=.01 candidate was selected before this ChineseEEG run-07 evaluation.",
            "No target-side representation, subject, item, model, lambda, layer or checkpoint selection is performed.",
            "ChineseEEG is a secondary consistency target, not a fresh independent confirmation, because it contributed to the broader NeuroSem development history.",
            "No rescue tuning is permitted after this result.",
        ],
    }
    (out / "summary.json").write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload["primary_secondary_result"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
