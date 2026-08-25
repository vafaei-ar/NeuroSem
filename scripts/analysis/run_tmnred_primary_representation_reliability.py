#!/usr/bin/env python3
import argparse, csv, json, math, unicodedata
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.io import loadmat
from scipy.signal import resample_poly
from scipy.spatial.distance import pdist, squareform
from scipy.stats import spearmanr

READY_SUBJECTS = [f"sub-{i:02d}" for i in range(1,31) if i != 25]
SESSIONS = [f"ses-{i}" for i in range(1,9)]
CANDIDATES = ["row_mean_all", "row_std_all", "relative_8bin_all"]


def as_list(v):
    if v is None: return []
    if isinstance(v, list): return v
    if hasattr(v, "tolist"):
        v = v.tolist()
        return v if isinstance(v, list) else [v]
    return [v]


def source_from_set(path):
    d = loadmat(path, simplify_cells=True)
    eeg = d.get("EEG")
    if isinstance(eeg, dict): return eeg, "EEG_struct"
    return d, "flat_top_level"


def epoch_map(source):
    out = {}
    for e in as_list(source.get("event")):
        if not isinstance(e, dict): continue
        try:
            b = int(e.get("bepoch")); ep = int(e.get("epoch"))
        except Exception:
            continue
        if b in out and out[b] != ep:
            raise RuntimeError(f"inconsistent event epoch mapping for bepoch={b}")
        out[b] = ep
    return out


def load_signal(root, subject, session):
    set_path = root / f"derivatives/preproc/{subject}/{session}/{subject}-{session}z.set"
    src, layout = source_from_set(set_path)
    nbchan, pnts, trials = int(src["nbchan"]), int(src["pnts"]), int(src["trials"])
    srate, xmin = float(src["srate"]), float(src["xmin"])
    data = src.get("data")
    if isinstance(data, np.ndarray):
        arr = np.asarray(data, dtype=float)
    elif not isinstance(data, str):
        arr = np.asarray(data, dtype=float)
    else:
        fdt = set_path.with_suffix(".fdt")
        if not fdt.exists():
            legacy = set_path.parent / data
            fdt = legacy if legacy.exists() else fdt
        raw = np.fromfile(fdt, dtype="<f4")
        expected = nbchan * pnts * trials
        if raw.size != expected:
            raise RuntimeError(f"FDT size mismatch {fdt}: {raw.size} != {expected}")
        arr = raw.reshape((nbchan, pnts, trials), order="F").astype(float)
    arr = np.squeeze(arr)
    if arr.shape == (trials, pnts, nbchan): arr = arr.transpose(2,1,0)
    elif arr.shape == (pnts, nbchan, trials): arr = arr.transpose(1,0,2)
    elif arr.shape == (nbchan, trials, pnts): arr = arr.transpose(0,2,1)
    if arr.shape != (nbchan, pnts, trials):
        raise RuntimeError(f"unexpected EEG shape {arr.shape}, expected {(nbchan,pnts,trials)}")
    emap = epoch_map(src)
    if len(emap) != trials:
        raise RuntimeError(f"event.bepoch mapping count {len(emap)} != trials {trials}")
    if abs(srate - 200.0) > 1e-6:
        if abs(srate - 500.0) > 1e-6: raise RuntimeError(f"unexpected sampling rate {srate}")
        arr = resample_poly(arr, 2, 5, axis=1)
        srate = 200.0
    times = xmin + np.arange(arr.shape[1]) / srate
    keep = (times >= 0.0) & (times < 2.0)
    if keep.sum() < 300: raise RuntimeError("too few post-onset samples")
    return arr[:, keep, :], emap


def features(arr, candidate):
    # output trials x features
    if candidate == "row_mean_all":
        return arr.mean(axis=1).T
    if candidate == "row_std_all":
        return arr.std(axis=1, ddof=0).T
    if candidate == "relative_8bin_all":
        bins = np.array_split(np.arange(arr.shape[1]), 8)
        return np.concatenate([arr[:, b, :].mean(axis=1).T for b in bins], axis=1)
    raise ValueError(candidate)


def zscore_cols(x):
    x = np.asarray(x, float)
    mu = x.mean(axis=0); sd = x.std(axis=0)
    good = sd > 1e-12
    if good.sum() < 2: raise RuntimeError("insufficient nonconstant features")
    return (x[:, good] - mu[good]) / sd[good]


def is_cjk(ch):
    o = ord(ch)
    return (0x3400 <= o <= 0x4DBF) or (0x4E00 <= o <= 0x9FFF) or (0xF900 <= o <= 0xFAFF)


def stimulus_blocks(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for si in range(1,9):
        ws = wb[f"Block{si}"]
        rows = list(ws.iter_rows(min_row=2, max_row=51, values_only=True))
        texts = [str(r[1] or "") for r in rows]
        if len(texts) != 50: raise RuntimeError(f"Block{si} expected 50 rows")
        out[f"ses-{si}"] = texts
    return out


def nuisance_matrix(texts):
    n = len(texts); pairs=[]; vals=[]
    cjk = [set(ch for ch in t if is_cjk(ch)) for t in texts]
    clen = [sum(is_cjk(ch) for ch in t) for t in texts]
    punct = [sum(unicodedata.category(ch).startswith("P") for ch in t) for t in texts]
    for i in range(n):
        for j in range(i+1,n):
            inter, union = len(cjk[i]&cjk[j]), len(cjk[i]|cjk[j])
            jac_dist = 1.0 - (inter/union if union else 1.0)
            pairs.append((i+1,j+1))
            vals.append([abs(i-j), abs(clen[i]-clen[j]), abs(punct[i]-punct[j]), jac_dist])
    X=np.asarray(vals,float)
    mu=X.mean(0); sd=X.std(0); good=sd>1e-12
    X=(X[:,good]-mu[good])/sd[good]
    return pairs, X


def residualize(y, X):
    A=np.column_stack([np.ones(len(y)), X])
    beta=np.linalg.lstsq(A,y,rcond=None)[0]
    return y-A@beta


def fisher_mean(rs):
    rs=np.asarray([r for r in rs if np.isfinite(r)],float)
    if not len(rs): return float("nan")
    z=np.arctanh(np.clip(rs,-0.999999,0.999999))
    return float(np.tanh(z.mean()))


def boot_ci(values, seed=20260825, nboot=10000):
    v=np.asarray([x for x in values if np.isfinite(x)],float)
    rng=np.random.default_rng(seed)
    means=np.empty(nboot)
    for b in range(nboot): means[b]=rng.choice(v,len(v),replace=True).mean()
    return [float(np.quantile(means,.025)), float(np.quantile(means,.975))]


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--data-root",default="data/raw/tmnred")
    ap.add_argument("--input-freeze",default="outputs/tmnred_representation_input_materialization/latest/summary.json")
    ap.add_argument("--output-dir",default="outputs/tmnred_primary_representation_reliability/latest")
    ap.add_argument("--min-reference-edge-subjects",type=int,default=18)
    args=ap.parse_args()
    root=Path(args.data_root); out=Path(args.output_dir); out.mkdir(parents=True,exist_ok=True)
    freeze=json.loads(Path(args.input_freeze).read_text())
    if freeze.get("ready_subjects_all_8_sessions") != READY_SUBJECTS: raise SystemExit("frozen subject cohort mismatch")
    if freeze.get("item_cohort_failures"): raise SystemExit("item cohort freeze is not clean")
    for s in SESSIONS:
        if freeze["item_coverage_by_session"][s]["n_core_items"] != 50: raise SystemExit(f"unexpected frozen core size for {s}")
    blocks=stimulus_blocks(root/"derivatives/source material/source material.xlsx")
    nuisance={s:nuisance_matrix(blocks[s]) for s in SESSIONS}
    edge_index={s:{p:k for k,p in enumerate(nuisance[s][0])} for s in SESSIONS}

    all_metrics=[]; subject_rows=[]; loader=[]
    for cand in CANDIDATES:
        session_rdms={}
        for s in SESSIONS:
            pairlist,_=nuisance[s]; E=len(pairlist)
            mat=np.full((len(READY_SUBJECTS),E),np.nan)
            for qi,sub in enumerate(READY_SUBJECTS):
                arr,emap=load_signal(root,sub,s)
                f=features(arr,cand)
                items=sorted(emap)
                epidx=[emap[b]-1 for b in items]
                xf=zscore_cols(f[epidx,:])
                D=squareform(pdist(xf,metric="correlation"))
                for a in range(len(items)):
                    for b in range(a+1,len(items)):
                        key=(items[a],items[b])
                        k=edge_index[s].get(key)
                        if k is not None: mat[qi,k]=D[a,b]
                if cand==CANDIDATES[0]: loader.append({"subject":sub,"session":s,"n_items":len(items),"n_samples":arr.shape[1]})
            session_rdms[s]=mat

        per_subject={sub:{"raw":[],"resid":[],"edges":[]} for sub in READY_SUBJECTS}
        pairwise_raw=[]; pairwise_resid=[]
        for s in SESSIONS:
            mat=session_rdms[s]; X=nuisance[s][1]
            for i,sub in enumerate(READY_SUBJECTS):
                others=np.delete(mat,i,axis=0)
                support=np.sum(np.isfinite(others),axis=0)
                ref=np.nanmean(others,axis=0)
                mask=np.isfinite(mat[i]) & np.isfinite(ref) & (support>=args.min_reference_edge_subjects)
                if mask.sum()<100: raise RuntimeError(f"too few LOO edges {cand} {sub} {s}: {mask.sum()}")
                y=mat[i,mask]; r=ref[mask]; Xm=X[mask]
                raw=float(spearmanr(y,r).statistic)
                resid=float(spearmanr(residualize(y,Xm),residualize(r,Xm)).statistic)
                per_subject[sub]["raw"].append(raw); per_subject[sub]["resid"].append(resid); per_subject[sub]["edges"].append(int(mask.sum()))
            for i in range(len(READY_SUBJECTS)):
                for j in range(i+1,len(READY_SUBJECTS)):
                    mask=np.isfinite(mat[i]) & np.isfinite(mat[j])
                    if mask.sum()<100: continue
                    pairwise_raw.append(float(spearmanr(mat[i,mask],mat[j,mask]).statistic))
                    pairwise_resid.append(float(spearmanr(residualize(mat[i,mask],X[mask]),residualize(mat[j,mask],X[mask])).statistic))

        agg_raw=[]; agg_res=[]
        for sub in READY_SUBJECTS:
            r=fisher_mean(per_subject[sub]["raw"]); q=fisher_mean(per_subject[sub]["resid"])
            agg_raw.append(r); agg_res.append(q)
            subject_rows.append({"candidate":cand,"subject":sub,"raw_loo":r,"resid_loo":q,"fraction_sessions_positive_resid":float(np.mean(np.asarray(per_subject[sub]["resid"])>0)),"mean_edges":float(np.mean(per_subject[sub]["edges"]))})
        all_metrics.append({
            "candidate":cand,
            "primary":cand=="row_mean_all",
            "mean_raw_loo":float(np.mean(agg_raw)),
            "mean_resid_loo":float(np.mean(agg_res)),
            "median_resid_loo":float(np.median(agg_res)),
            "resid_loo_bootstrap_95ci":boot_ci(agg_res),
            "fraction_subjects_positive_resid":float(np.mean(np.asarray(agg_res)>0)),
            "mean_raw_pairwise":float(np.nanmean(pairwise_raw)),
            "mean_resid_pairwise":float(np.nanmean(pairwise_resid)),
        })

    with (out/"candidate_metrics.csv").open("w",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=["candidate","primary","mean_raw_loo","mean_resid_loo","median_resid_loo","ci_low","ci_high","fraction_subjects_positive_resid","mean_raw_pairwise","mean_resid_pairwise"]); w.writeheader()
        for m in all_metrics:
            w.writerow({**{k:m[k] for k in ["candidate","primary","mean_raw_loo","mean_resid_loo","median_resid_loo","fraction_subjects_positive_resid","mean_raw_pairwise","mean_resid_pairwise"]},"ci_low":m["resid_loo_bootstrap_95ci"][0],"ci_high":m["resid_loo_bootstrap_95ci"][1]})
    with (out/"subject_metrics.csv").open("w",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=list(subject_rows[0])); w.writeheader(); w.writerows(subject_rows)
    payload={
        "schema_version":1,"dataset":"TMNRED","model_blind":True,
        "frozen_subjects":READY_SUBJECTS,"excluded_subject":"sub-25","resampled_subject":"sub-23",
        "window_seconds":[0.0,2.0],"target_sampling_hz":200,
        "primary_candidate":"row_mean_all","sensitivity_candidates":["row_std_all","relative_8bin_all"],
        "rdm":"feature-wise z-score across items, correlation distance",
        "nuisance_rdms":["absolute trial-position difference","CJK character-count difference","punctuation-count difference","CJK character-set Jaccard distance"],
        "min_reference_edge_subjects":args.min_reference_edge_subjects,
        "participant_aggregation":"Fisher-z mean across eight within-session LOO Spearman reliabilities",
        "metrics":all_metrics,
        "loader_summary":loader,
        "guardrails":["No language-model embeddings are loaded.","No candidate or time-window selection is performed from TMNRED outcomes.","The all-sensor temporal mean is the prospectively designated primary external replication endpoint."],
    }
    (out/"summary.json").write_text(json.dumps(payload,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"status":"ok","primary":all_metrics[0],"output_dir":str(out)},indent=2))

if __name__=="__main__": main()
