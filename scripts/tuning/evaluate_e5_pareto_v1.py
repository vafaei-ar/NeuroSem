#!/usr/bin/env python3
"""Exploratory E5 neural-alignment / semantic-utility Pareto evaluation.

This script evaluates a prespecified neural-loss-weight grid after training. Run-07 and
C-MTEB STS are reused only as exploratory/mechanistic outcomes because both were already
observed before this dose-response experiment.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl
from scipy.stats import spearmanr

MODEL_ID = "intfloat/multilingual-e5-large"
MODEL_REVISION = "3d7cfbdacd47fdda877c5cd8a79fbcc4f2a574f3"
PREFIX = "query: "
SUBJECTS = ["sub-04", "sub-05", "sub-06", "sub-07", "sub-08", "sub-09", "sub-10", "sub-13", "sub-14", "sub-15"]

GRID = [0.00, 0.01, 0.03, 0.10, 0.30, 1.00]
LABELS = {
    0.00: "lambda_0",
    0.01: "lambda_0p01",
    0.03: "lambda_0p03",
    0.10: "lambda_0p10",
    0.30: "lambda_0p30",
    1.00: "lambda_1",
}
ANCHOR_ADAPTERS = {
    0.00: Path("outputs/e5_neural_tuning_v1/text_only/20260823_181507/adapter"),
    1.00: Path("outputs/e5_neural_tuning_v1/neural/20260823_181609/adapter"),
}
ANCHOR_SUMMARIES = {
    0.00: Path("outputs/e5_neural_tuning_v1/text_only/20260823_181507/summary.json"),
    1.00: Path("outputs/e5_neural_tuning_v1/neural/20260823_181609/summary.json"),
}

TASKS = {
    "AFQMC": {"repo": "mteb/AFQMC", "config": None, "revision": "2fbe82b59f6587c54f8af3b6f0716f2ea5a19ced", "split": "validation"},
    "ATEC": {"repo": "mteb/ATEC", "config": None, "revision": "a34edc2663ba1dc46c283ffb5e8303e37eaa1ff6", "split": "test"},
    "BQ": {"repo": "mteb/BQ", "config": None, "revision": "3b4fd209dbf361fa5be8344772729b93ca6af5b7", "split": "test"},
    "LCQMC": {"repo": "mteb/LCQMC", "config": None, "revision": "20142739d5bd9f501c3668f3faef85a58f26c8e5", "split": "test"},
    "PAWSX": {"repo": "mteb/PAWSX", "config": None, "revision": "bd129d4230ee0551b5469c566bced8da75abae0a", "split": "test"},
    "QBQTC": {"repo": "mteb/QBQTC", "config": None, "revision": "8bc4e5da5dddcd334f22853274642fe6139544d8", "split": "test"},
    "STS22 (zh)": {"repo": "mteb/sts22-crosslingual-sts", "config": "zh", "revision": "91d97a5b9d761e285ac3e1b4f239797bbd21c4b8", "split": "test"},
    "STSB": {"repo": "mteb/STSB", "config": None, "revision": "e036defee8f190911de6fc825a944b928c8cda53", "split": "test"},
}


def latest_dir(root: Path, required: str) -> Path:
    candidates = [d for d in root.iterdir() if d.is_dir() and (d / required).exists()] if root.exists() else []
    if not candidates:
        raise FileNotFoundError(f"No directory containing {required} under {root}")
    return sorted(candidates)[-1]


def intermediate_root(lambda_value: float, tuning_root: Path) -> Path:
    return tuning_root / LABELS[lambda_value]


def adapter_for(lambda_value: float, tuning_root: Path) -> Path:
    if lambda_value in ANCHOR_ADAPTERS:
        path = ANCHOR_ADAPTERS[lambda_value]
    else:
        path = latest_dir(intermediate_root(lambda_value, tuning_root) / "neural", "summary.json") / "adapter"
    if not path.exists():
        raise FileNotFoundError(f"Missing adapter for lambda={lambda_value}: {path}")
    return path


def tuning_summary_for(lambda_value: float, tuning_root: Path) -> Path:
    if lambda_value in ANCHOR_SUMMARIES:
        path = ANCHOR_SUMMARIES[lambda_value]
    else:
        path = latest_dir(intermediate_root(lambda_value, tuning_root) / "neural", "summary.json") / "summary.json"
    if not path.exists():
        raise FileNotFoundError(f"Missing tuning summary for lambda={lambda_value}: {path}")
    return path


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    return (hidden * mask).sum(dim=1) / mask.sum(dim=1).clamp_min(1.0)


def encode(model, tokenizer, texts, device, batch_size, max_length):
    import torch

    chunks = []
    with torch.inference_mode():
        for start in range(0, len(texts), batch_size):
            batch = [PREFIX + str(x) for x in texts[start:start + batch_size]]
            enc = tokenizer(batch, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
            attention = enc["attention_mask"]
            enc = {k: v.to(device) for k, v in enc.items()}
            out = model(**enc, return_dict=True)
            pooled = masked_mean(out.last_hidden_state, attention.to(device).bool())
            pooled = torch.nn.functional.normalize(pooled, p=2, dim=1)
            chunks.append(pooled.cpu().numpy().astype(np.float32))
    return np.concatenate(chunks, axis=0)


def load_model(adapter_path: Path, device: str):
    from peft import PeftModel
    from transformers import AutoModel, AutoTokenizer

    tokenizer = AutoTokenizer.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    base = AutoModel.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    model = PeftModel.from_pretrained(base, adapter_path)
    model.eval().to(device)
    return tokenizer, model


def load_task(spec):
    from datasets import load_dataset

    kwargs = {
        "path": spec["repo"],
        "split": spec["split"],
        "revision": spec["revision"],
    }
    if spec["config"] is not None:
        kwargs["name"] = spec["config"]
    ds = load_dataset(**kwargs)
    required = {"sentence1", "sentence2", "score"}
    if not required.issubset(ds.column_names):
        raise RuntimeError(f"Frozen task lacks required columns: {spec['repo']}")
    gold = np.asarray(ds["score"], dtype=np.float64)
    if not gold.size or not np.isfinite(gold).all() or np.unique(gold).size < 2:
        raise RuntimeError(f"Frozen task has unusable gold labels: {spec['repo']}")
    return ds


def run_rsa(embedding_root: Path, feature_root: Path, output_root: Path, permutations: int, workers: int, chunk_size: int):
    cmd = [
        sys.executable,
        "scripts/analysis/assess_chineseeeg_run07_holdout_fast.py",
        "--embedding-root", str(embedding_root),
        "--feature-root", str(feature_root),
        "--subjects", *SUBJECTS,
        "--permutations", str(permutations),
        "--workers", str(workers),
        "--chunk-size", str(chunk_size),
        "--output-dir", str(output_root),
    ]
    print("+", " ".join(cmd), flush=True)
    subprocess.run(cmd, check=True)
    return latest_dir(output_root, "summary.json") / "summary.json"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--tuning-root", type=Path, default=Path("outputs/e5_neural_tuning_pareto_v1"))
    parser.add_argument("--feature-root", type=Path, default=Path("outputs/chineseeeg_row_features/run-07"))
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/e5_pareto_v1"))
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--max-length", type=int, default=64)
    parser.add_argument("--permutations", type=int, default=10000)
    parser.add_argument("--workers", type=int, default=16)
    parser.add_argument("--chunk-size", type=int, default=50)
    parser.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    args = parser.parse_args()

    import datasets
    import peft
    import torch
    import transformers

    device = "cuda" if args.device == "auto" and torch.cuda.is_available() else args.device
    if args.device == "auto" and not torch.cuda.is_available():
        device = "cpu"
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("CUDA requested but unavailable")

    dataset = args.dataset.expanduser().resolve()
    workbook = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / "segmented_Chinense_novel_run_7.xlsx"
    if not workbook.exists():
        raise SystemExit(f"Run-07 workbook not materialized: {workbook}")
    run07_texts = read_rows(workbook)
    if not run07_texts:
        raise SystemExit("No run-07 text rows found")

    loaded_tasks = {name: load_task(spec) for name, spec in TASKS.items()}

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    root = (args.output_dir / stamp).resolve()
    root.mkdir(parents=True, exist_ok=False)

    points = []
    for lambda_value in GRID:
        label = LABELS[lambda_value]
        print(f"\n=== Pareto point {label} (lambda={lambda_value}) ===", flush=True)
        adapter = adapter_for(lambda_value, args.tuning_root)
        tuning_summary_path = tuning_summary_for(lambda_value, args.tuning_root)
        tuning_summary = json.loads(tuning_summary_path.read_text(encoding="utf-8"))
        run06_corr = float(tuning_summary["final_run06_neural_corr"])

        tokenizer, model = load_model(adapter, device)

        emb = encode(model, tokenizer, run07_texts, device, args.batch_size, args.max_length)
        emb_dir = root / "run07_embeddings" / label
        emb_dir.mkdir(parents=True, exist_ok=False)
        np.save(emb_dir / "embeddings.npy", emb)
        (emb_dir / "texts.json").write_text(json.dumps(run07_texts, ensure_ascii=False, indent=2), encoding="utf-8")
        emb_summary = {
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "arm": label,
            "lambda": lambda_value,
            "model_id": MODEL_ID,
            "model_revision": MODEL_REVISION,
            "adapter_dir": str(adapter.resolve()),
            "run_number": 7,
            "n_rows": len(run07_texts),
            "embedding_shape": list(emb.shape),
            "pooling": "attention-mask mean, L2 normalized",
            "input_prefix": PREFIX,
            "max_length": args.max_length,
            "device": device,
            "evaluation_only": True,
            "exploratory_reuse": True,
        }
        (emb_dir / "summary.json").write_text(json.dumps(emb_summary, indent=2), encoding="utf-8")

        rsa_summary_path = run_rsa(
            emb_dir.parent / label,
            args.feature_root,
            root / "run07_rsa" / label,
            args.permutations,
            args.workers,
            args.chunk_size,
        )
        rsa_summary = json.loads(rsa_summary_path.read_text(encoding="utf-8"))
        run07_mean = float(rsa_summary["observed"]["mean"])

        task_scores = {}
        for task, ds in loaded_tasks.items():
            s1 = [str(x) for x in ds["sentence1"]]
            s2 = [str(x) for x in ds["sentence2"]]
            gold = np.asarray(ds["score"], dtype=np.float64)
            e1 = encode(model, tokenizer, s1, device, args.batch_size, args.max_length)
            e2 = encode(model, tokenizer, s2, device, args.batch_size, args.max_length)
            sim = np.einsum("ij,ij->i", e1, e2, dtype=np.float64)
            rho = float(spearmanr(sim, gold).statistic)
            if not np.isfinite(rho):
                raise RuntimeError(f"Non-finite Spearman for lambda={lambda_value}/{task}")
            task_scores[task] = rho
            print(f"  {task}: {rho:.6f}", flush=True)
        sts_mean = float(np.mean(list(task_scores.values())))

        points.append({
            "lambda": lambda_value,
            "label": label,
            "adapter_path": str(adapter.resolve()),
            "tuning_summary_path": str(tuning_summary_path.resolve()),
            "run06_neural_corr": run06_corr,
            "run07_neural_rsa_mean": run07_mean,
            "run07_neural_rsa_median": float(rsa_summary["observed"]["median"]),
            "run07_neural_rsa_p": float(rsa_summary["inference"]["p_ge_observed"]),
            "external_sts_mean": sts_mean,
            "external_sts_task_scores": task_scores,
        })
        print(
            f"lambda={lambda_value:.2f} | run06={run06_corr:.6f} | "
            f"run07={run07_mean:.6f} | external_STS={sts_mean:.6f}",
            flush=True,
        )

        del model
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    anchor = next(p for p in points if p["lambda"] == 0.0)
    for p in points:
        p["delta_run07_vs_lambda0"] = p["run07_neural_rsa_mean"] - anchor["run07_neural_rsa_mean"]
        p["delta_external_sts_vs_lambda0"] = p["external_sts_mean"] - anchor["external_sts_mean"]

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "protocol": "docs/e5_pareto_exploratory_protocol_v1.md",
        "analysis_status": "exploratory/mechanistic",
        "model_id": MODEL_ID,
        "model_revision": MODEL_REVISION,
        "grid": GRID,
        "anchor_policy": "lambda=0 and lambda=1 reuse exact completed frozen E5 adapters; intermediate lambdas are newly trained",
        "run07_reuse_warning": "Run-07 was observed before this dose-response experiment and is descriptive only.",
        "external_sts_reuse_warning": "The eight-task STS endpoint was observed before this dose-response experiment and is descriptive only.",
        "dataset_provenance": {task: {k: v for k, v in spec.items() if k in {"repo", "config", "revision", "split"}} for task, spec in TASKS.items()},
        "software": {
            "transformers": transformers.__version__,
            "datasets": datasets.__version__,
            "peft": peft.__version__,
        },
        "points": points,
    }
    summary_path = root / "combined_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    csv_path = root / "pareto_points.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "lambda", "run06_neural_corr", "run07_neural_rsa_mean", "external_sts_mean",
                "delta_run07_vs_lambda0", "delta_external_sts_vs_lambda0",
            ],
        )
        writer.writeheader()
        for p in points:
            writer.writerow({k: p[k] for k in writer.fieldnames})

    latest = args.output_dir / "latest"
    latest.mkdir(parents=True, exist_ok=True)
    shutil.copy2(summary_path, latest / "combined_summary.json")
    shutil.copy2(csv_path, latest / "pareto_points.csv")

    print("\n=== E5 exploratory Pareto summary ===")
    for p in points:
        print(
            f"lambda={p['lambda']:.2f} | run06={p['run06_neural_corr']:.6f} | "
            f"run07={p['run07_neural_rsa_mean']:.6f} | STS={p['external_sts_mean']:.6f} | "
            f"dRun07={p['delta_run07_vs_lambda0']:+.6f} | dSTS={p['delta_external_sts_vs_lambda0']:+.6f}"
        )
    print(f"Summary: {latest / 'combined_summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
