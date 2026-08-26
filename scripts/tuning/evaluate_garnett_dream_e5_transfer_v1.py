#!/usr/bin/env python3
"""Frozen ChineseEEG-to-Garnett-Dream multilingual-E5 transfer test.

Primary confirmatory contrast: ChineseEEG-trained lambda=0.10 neural-guided adapter
versus lambda=0 text-only adapter, evaluated against the prospectively frozen Garnett
Dream row_mean_all EEG geometry. Garnett Dream is same-participant/new-text validation.
No Garnett tuning or outcome-driven representation/model selection is performed here.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.spatial.distance import pdist
from scipy.stats import spearmanr

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.analysis.run_garnett_dream_primary_reliability import (
    bootstrap_ci,
    companion_vhdr_by_run,
    exact_signflip,
    features_for_run,
    fisher_mean,
    rdm_from_features,
    read_csv,
    residualize,
)
from scripts.tuning.evaluate_tmnred_e5_transfer_v1 import (
    LAMBDA_010_ROOT,
    TEXT_ONLY_ADAPTER,
    encode_texts,
    latest_completed_adapter,
    load_adapter,
)

ARMS = ["lambda_0", "lambda_0p10"]
EXPECTED_HEADER = "Chinese_text"


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def safe_rho(a: np.ndarray, b: np.ndarray) -> float:
    r = float(spearmanr(a, b).statistic)
    if not np.isfinite(r):
        raise RuntimeError("non-finite Spearman RSA")
    return r


def is_cjk(ch: str) -> bool:
    o = ord(ch)
    return (0x3400 <= o <= 0x4DBF) or (0x4E00 <= o <= 0x9FFF) or (0xF900 <= o <= 0xFAFF)


def load_text_rows(path: Path, expected_n: int) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if len(wb.worksheets) != 1:
        wb.close()
        raise RuntimeError(f"expected one worksheet in {path}, found {len(wb.worksheets)}")
    ws = wb.worksheets[0]
    nonempty = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if vals:
            nonempty.append(vals)
    wb.close()
    if not nonempty or len(nonempty[0]) != 1 or nonempty[0][0] != EXPECTED_HEADER:
        raise RuntimeError(f"unexpected Garnett XLSX header in {path}")
    texts = []
    for i, vals in enumerate(nonempty[1:], start=2):
        if len(vals) != 1:
            raise RuntimeError(f"expected one text cell at physical row {i} in {path}, got {len(vals)}")
        texts.append(vals[0])
    if len(texts) != expected_n or any(not t for t in texts):
        raise RuntimeError(f"text count mismatch in {path}: {len(texts)} != {expected_n}")
    return texts


def load_frozen_texts(mapping_summary: dict, data_root: Path, expected_counts: dict[int, int]) -> tuple[dict[int, list[str]], dict[int, str]]:
    gate = mapping_summary.get("freeze_gate", {})
    if not gate.get("exact_row_text_mapping_identified") or not gate.get("ready_to_freeze_model_validation_text_mapping"):
        raise RuntimeError("Garnett exact row-text mapping is not freeze-ready")
    matches = mapping_summary.get("exact_run_matches", [])
    if len(matches) != 18:
        raise RuntimeError(f"expected 18 exact Garnett workbook matches, found {len(matches)}")
    by_chapter = {}
    provenance = {}
    for rec in matches:
        chapter = int(rec["chapter"])
        if chapter not in expected_counts:
            raise RuntimeError(f"unexpected mapped chapter {chapter}")
        if int(rec["expected_items"]) != expected_counts[chapter]:
            raise RuntimeError(f"mapping/item count mismatch for chapter {chapter}")
        path = data_root / rec["path"]
        texts = load_text_rows(path, expected_counts[chapter])
        by_chapter[chapter] = texts
        provenance[chapter] = rec["path"]
    if sorted(by_chapter) != list(range(1, 19)):
        raise RuntimeError("frozen Garnett text map does not cover chapters 1..18")
    return by_chapter, provenance


def text_nuisance_edges(texts: list[str]) -> dict[str, np.ndarray]:
    n = len(texts)
    order = pdist(np.arange(1, n + 1, dtype=float)[:, None], metric="cityblock")
    cjk_sets = [set(ch for ch in t if is_cjk(ch)) for t in texts]
    char_counts = np.asarray([sum(is_cjk(ch) for ch in t) for t in texts], dtype=float)
    punct_counts = np.asarray([sum(unicodedata.category(ch).startswith("P") for ch in t) for t in texts], dtype=float)
    char_diff = pdist(char_counts[:, None], metric="cityblock")
    punct_diff = pdist(punct_counts[:, None], metric="cityblock")
    jaccard = []
    for i in range(n):
        for j in range(i + 1, n):
            union = len(cjk_sets[i] | cjk_sets[j])
            inter = len(cjk_sets[i] & cjk_sets[j])
            jaccard.append(1.0 - (inter / union if union else 1.0))
    jaccard = np.asarray(jaccard, dtype=float)
    for name, arr in {
        "order": order,
        "character_count": char_diff,
        "punctuation_count": punct_diff,
        "character_set_jaccard": jaccard,
    }.items():
        if not np.isfinite(arr).all() or len(arr) != n * (n - 1) // 2:
            raise RuntimeError(f"invalid nuisance RDM {name}")
    return {
        "order": order,
        "character_count": char_diff,
        "punctuation_count": punct_diff,
        "character_set_jaccard": jaccard,
    }


def build_model_edges(texts_by_chapter: dict[int, list[str]], device: str):
    import torch

    specs = {
        "lambda_0": TEXT_ONLY_ADAPTER,
        "lambda_0p10": latest_completed_adapter(LAMBDA_010_ROOT),
    }
    chapters = list(range(1, 19))
    flat = [t for ch in chapters for t in texts_by_chapter[ch]]
    counts = {ch: len(texts_by_chapter[ch]) for ch in chapters}
    out = {}
    provenance = {}
    for label, adapter in specs.items():
        print(f"Loading model arm: {label}", flush=True)
        tok, model = load_adapter(adapter, device)
        emb = encode_texts(model, tok, flat, device)
        if emb.shape[0] != len(flat):
            raise RuntimeError(f"expected {len(flat)} Garnett embeddings, got {emb.shape}")
        out[label] = {}
        off = 0
        for ch in chapters:
            n = counts[ch]
            e = emb[off : off + n]
            d = pdist(e, metric="cosine")
            if not np.isfinite(d).all():
                raise RuntimeError(f"non-finite model RDM for {label} chapter {ch}")
            out[label][ch] = d
            off += n
        provenance[label] = str(adapter.resolve())
        del model
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
    return out, provenance


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-root", type=Path, default=Path("data/raw/chineseeeg"))
    ap.add_argument("--input-freeze", type=Path, default=Path("outputs/garnett_dream_input_materialization/latest/summary.json"))
    ap.add_argument("--session-inventory", type=Path, default=Path("outputs/garnett_dream_input_materialization/latest/session_inventory.csv"))
    ap.add_argument("--item-identity", type=Path, default=Path("outputs/garnett_dream_input_materialization/latest/item_identity.csv"))
    ap.add_argument("--reliability-summary", type=Path, default=Path("outputs/garnett_dream_primary_reliability/latest/summary.json"))
    ap.add_argument("--mapping-freeze", type=Path, default=Path("outputs/garnett_dream_segmented_xlsx_mapping_probe_v1/latest/summary.json"))
    ap.add_argument("--output-dir", type=Path, default=Path("outputs/garnett_dream_e5_transfer_v1/latest"))
    ap.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    args = ap.parse_args()

    import torch

    if args.device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    else:
        device = args.device
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("CUDA requested but unavailable")

    data_root = args.data_root.resolve()
    freeze = json.loads(args.input_freeze.read_text(encoding="utf-8"))
    if not freeze.get("freeze_gate", {}).get("ready_for_reliability"):
        raise SystemExit("Garnett structural freeze is not ready")
    if freeze.get("n_ready_subjects") != 10 or freeze.get("n_ready_runs") != 171:
        raise SystemExit("unexpected frozen Garnett cohort/run count")
    expected_counts = {int(k): int(v) for k, v in freeze.get("chapter_item_counts", {}).items()}
    if sorted(expected_counts) != list(range(1, 19)) or sum(expected_counts.values()) != 9047:
        raise SystemExit("unexpected frozen Garnett chapter/item counts")

    reliability = json.loads(args.reliability_summary.read_text(encoding="utf-8"))
    primary_rel = reliability.get("summaries", {}).get("row_mean_all", {})
    ci = primary_rel.get("participant_bootstrap_95ci_residual_mean", [])
    if len(ci) != 2 or float(ci[0]) <= 0 or float(primary_rel.get("mean_residual_loo", 0.0)) <= 0:
        raise SystemExit("prospectively required Garnett EEG reliability gate is not positive")

    mapping = json.loads(args.mapping_freeze.read_text(encoding="utf-8"))
    texts_by_chapter, text_provenance = load_frozen_texts(mapping, data_root, expected_counts)
    text_nuisance = {ch: text_nuisance_edges(texts_by_chapter[ch]) for ch in range(1, 19)}
    model_edges, model_provenance = build_model_edges(texts_by_chapter, device)

    inventory = read_csv(args.session_inventory)
    items = read_csv(args.item_identity)
    vhdrs = companion_vhdr_by_run(inventory)
    items_by_key = defaultdict(list)
    for r in items:
        items_by_key[(str(r["subject"]), int(r["run"]), int(r["chapter"]))].append(r)
    if len(vhdrs) != 171:
        raise SystemExit(f"expected 171 frozen Garnett EEG runs, found {len(vhdrs)}")

    chapter_rows = []
    by_subject = defaultdict(lambda: {a: [] for a in ARMS})
    chapters_by_subject = defaultdict(list)

    for key, vhdr_rel in sorted(vhdrs.items(), key=lambda kv: (kv[0][2], kv[0][0])):
        sub, _, chapter = key
        feats, durations, _ = features_for_run(data_root, vhdr_rel, items_by_key[key])
        x = feats["row_mean_all"]
        n = expected_counts[chapter]
        if x.shape[0] != n or len(durations) != n:
            raise RuntimeError(f"frozen item count mismatch for {sub} chapter {chapter}")
        neural = rdm_from_features(x)
        duration_rdm = pdist(np.asarray(durations, dtype=float)[:, None], metric="cityblock")
        tn = text_nuisance[chapter]
        nuisances = [
            tn["order"],
            duration_rdm,
            tn["character_count"],
            tn["punctuation_count"],
            tn["character_set_jaccard"],
        ]
        neural_resid = residualize(neural, nuisances)
        vals = {}
        for arm in ARMS:
            model_resid = residualize(model_edges[arm][chapter], nuisances)
            rho = safe_rho(neural_resid, model_resid)
            vals[arm] = rho
            by_subject[sub][arm].append(rho)
        chapters_by_subject[sub].append(chapter)
        chapter_rows.append({
            "subject": sub,
            "chapter": chapter,
            "n_items": n,
            "lambda_0_resid_rsa": vals["lambda_0"],
            "lambda_0p10_resid_rsa": vals["lambda_0p10"],
            "delta_0p10_minus_0": vals["lambda_0p10"] - vals["lambda_0"],
            "n_edges": len(neural),
        })

    subjects = sorted(by_subject)
    if len(subjects) != 10:
        raise SystemExit(f"expected 10 Garnett subjects, found {len(subjects)}")

    subject_rows = []
    diffs = []
    for sub in subjects:
        a0 = fisher_mean(by_subject[sub]["lambda_0"])
        a1 = fisher_mean(by_subject[sub]["lambda_0p10"])
        d = a1 - a0
        diffs.append(d)
        subject_rows.append({
            "subject": sub,
            "n_chapters": len(chapters_by_subject[sub]),
            "chapters": ",".join(f"CH{x:02d}" for x in sorted(chapters_by_subject[sub])),
            "lambda_0_resid_rsa": a0,
            "lambda_0p10_resid_rsa": a1,
            "delta_0p10_minus_0": d,
        })

    diffs_arr = np.asarray(diffs, dtype=float)
    primary = {
        "contrast": "lambda_0p10_minus_lambda_0",
        "mean_delta": float(diffs_arr.mean()),
        "median_delta": float(np.median(diffs_arr)),
        "fraction_subjects_positive": float(np.mean(diffs_arr > 0)),
        "participant_bootstrap_95ci_mean_delta": bootstrap_ci(diffs),
        "exact_signflip": exact_signflip(diffs),
    }

    out = args.output_dir.resolve()
    out.mkdir(parents=True, exist_ok=True)
    write_csv(out / "subject_results.csv", subject_rows)
    write_csv(out / "chapter_results.csv", chapter_rows)
    payload = {
        "schema_version": 1,
        "dataset": "ChineseEEG Garnett Dream",
        "analysis": "single frozen multilingual-E5 same-participant/new-text transfer test",
        "primary_eeg_representation": "row_mean_all",
        "primary_contrast": "ChineseEEG-trained E5 lambda=0.10 neural-guided minus lambda=0 text-only",
        "no_garnett_tuning": True,
        "n_frozen_subjects": len(subjects),
        "n_frozen_subject_runs": len(chapter_rows),
        "n_unique_text_items_across_18_chapters": int(sum(expected_counts.values())),
        "text_mapping_rule": mapping.get("freeze_gate", {}).get("mapping_rule"),
        "text_workbook_provenance": {str(k): v for k, v in sorted(text_provenance.items())},
        "nuisance_rdms": [
            "within-chapter row/order difference",
            "participant-specific presentation-duration difference",
            "Chinese-character-count difference",
            "Unicode punctuation-count difference",
            "Chinese-character-set Jaccard distance",
        ],
        "participant_aggregation": "equal-weight Fisher-z mean across available chapter-specific nuisance-residualized Spearman RSAs",
        "model_provenance": model_provenance,
        "reliability_gate": {
            "mean_residual_loo": primary_rel.get("mean_residual_loo"),
            "bootstrap_95ci": ci,
            "passed_before_model_outcome": True,
        },
        "primary_result": primary,
        "guardrails": [
            "Garnett EEG reliability was established before this model-transfer test.",
            "Exact segmented-XLSX text mapping was frozen model-blind before model evaluation.",
            "No Garnett representation, participant, chapter, item, sensor, window, model, or lambda selection is performed from transfer outcomes.",
            "The sole confirmatory neural-guidance contrast is lambda=0.10 versus lambda=0 on row_mean_all.",
            "Garnett Dream is same-participant/new-text validation, not independent-cohort replication.",
        ],
    }
    (out / "summary.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"status": "ok", "primary_result": primary, "output_dir": str(out)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
