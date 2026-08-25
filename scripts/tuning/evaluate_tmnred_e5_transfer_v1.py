#!/usr/bin/env python3
"""Frozen TMNRED external-transfer evaluation for NeuroSem E5.

Primary confirmatory contrast: ChineseEEG-trained lambda=0.10 neural-guided adapter
versus the ChineseEEG-trained lambda=0 text-only adapter, evaluated against the
prospectively frozen TMNRED all-sensor temporal-mean EEG geometry.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.io import loadmat
from scipy.signal import resample_poly
from scipy.spatial.distance import pdist, squareform
from scipy.stats import spearmanr

MODEL_ID = "intfloat/multilingual-e5-large"
MODEL_REVISION = "3d7cfbdacd47fdda877c5cd8a79fbcc4f2a574f3"
PREFIX = "query: "
READY_SUBJECTS = [f"sub-{i:02d}" for i in range(1, 31) if i != 25]
SESSIONS = [f"ses-{i}" for i in range(1, 9)]
TEXT_ONLY_ADAPTER = Path("outputs/e5_neural_tuning_v1/text_only/20260823_181507/adapter")
LAMBDA_1_ADAPTER = Path("outputs/e5_neural_tuning_v1/neural/20260823_181609/adapter")
LAMBDA_010_ROOT = Path("outputs/e5_neural_tuning_pareto_v1/lambda_0p10/neural")


def as_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return v
    if hasattr(v, "tolist"):
        v = v.tolist()
        return v if isinstance(v, list) else [v]
    return [v]


def source_from_set(path: Path):
    d = loadmat(path, simplify_cells=True)
    eeg = d.get("EEG")
    if isinstance(eeg, dict):
        return eeg
    return d


def epoch_map(source: dict) -> dict[int, int]:
    out = {}
    for e in as_list(source.get("event")):
        if not isinstance(e, dict):
            continue
        try:
            b = int(e.get("bepoch"))
            ep = int(e.get("epoch"))
        except Exception:
            continue
        if b in out and out[b] != ep:
            raise RuntimeError(f"inconsistent event epoch mapping for bepoch={b}")
        out[b] = ep
    return out


def load_signal(root: Path, subject: str, session: str):
    set_path = root / f"derivatives/preproc/{subject}/{session}/{subject}-{session}z.set"
    src = source_from_set(set_path)
    nbchan = int(src["nbchan"])
    pnts = int(src["pnts"])
    trials = int(src["trials"])
    srate = float(src["srate"])
    xmin = float(src["xmin"])
    data = src.get("data")

    if isinstance(data, np.ndarray):
        arr = np.asarray(data, dtype=float)
    elif not isinstance(data, str):
        arr = np.asarray(data, dtype=float)
    else:
        fdt = set_path.with_suffix(".fdt")
        if not fdt.exists():
            legacy = set_path.parent / data
            fdt = legacy if legacy.exists() else fdt
        raw = np.fromfile(fdt, dtype="<f4")
        expected = nbchan * pnts * trials
        if raw.size != expected:
            raise RuntimeError(f"FDT size mismatch {fdt}: {raw.size} != {expected}")
        arr = raw.reshape((nbchan, pnts, trials), order="F").astype(float)

    arr = np.squeeze(arr)
    if arr.shape == (trials, pnts, nbchan):
        arr = arr.transpose(2, 1, 0)
    elif arr.shape == (pnts, nbchan, trials):
        arr = arr.transpose(1, 0, 2)
    elif arr.shape == (nbchan, trials, pnts):
        arr = arr.transpose(0, 2, 1)
    if arr.shape != (nbchan, pnts, trials):
        raise RuntimeError(f"unexpected EEG shape {arr.shape}, expected {(nbchan, pnts, trials)}")

    emap = epoch_map(src)
    if len(emap) != trials:
        raise RuntimeError(f"event.bepoch mapping count {len(emap)} != trials {trials}")

    if abs(srate - 200.0) > 1e-6:
        if abs(srate - 500.0) > 1e-6:
            raise RuntimeError(f"unexpected sampling rate {srate}")
        arr = resample_poly(arr, 2, 5, axis=1)
        srate = 200.0

    times = xmin + np.arange(arr.shape[1]) / srate
    keep = (times >= 0.0) & (times < 2.0)
    if keep.sum() < 300:
        raise RuntimeError("too few post-onset samples")
    return arr[:, keep, :], emap


def row_mean_features(arr: np.ndarray) -> np.ndarray:
    return arr.mean(axis=1).T


def zscore_cols(x: np.ndarray) -> np.ndarray:
    x = np.asarray(x, float)
    mu = x.mean(axis=0)
    sd = x.std(axis=0)
    good = sd > 1e-12
    if good.sum() < 2:
        raise RuntimeError("insufficient nonconstant EEG features")
    return (x[:, good] - mu[good]) / sd[good]


def is_cjk(ch: str) -> bool:
    o = ord(ch)
    return (0x3400 <= o <= 0x4DBF) or (0x4E00 <= o <= 0x9FFF) or (0xF900 <= o <= 0xFAFF)


def stimulus_blocks(path: Path) -> dict[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for si in range(1, 9):
        ws = wb[f"Block{si}"]
        rows = list(ws.iter_rows(min_row=2, max_row=51, values_only=True))
        texts = [str(r[1] or "") for r in rows]
        if len(texts) != 50 or any(not t for t in texts):
            raise RuntimeError(f"Block{si} expected 50 nonempty Chinese sentences")
        out[f"ses-{si}"] = texts
    return out


def nuisance_for_items(items: list[int], texts: list[str]) -> np.ndarray:
    vals = []
    cjk_sets = {i: set(ch for ch in texts[i - 1] if is_cjk(ch)) for i in items}
    clen = {i: sum(is_cjk(ch) for ch in texts[i - 1]) for i in items}
    punct = {i: sum(unicodedata.category(ch).startswith("P") for ch in texts[i - 1]) for i in items}
    for ai in range(len(items)):
        i = items[ai]
        for bj in range(ai + 1, len(items)):
            j = items[bj]
            inter = len(cjk_sets[i] & cjk_sets[j])
            union = len(cjk_sets[i] | cjk_sets[j])
            jac_dist = 1.0 - (inter / union if union else 1.0)
            vals.append([abs(i - j), abs(clen[i] - clen[j]), abs(punct[i] - punct[j]), jac_dist])
    X = np.asarray(vals, float)
    mu = X.mean(axis=0)
    sd = X.std(axis=0)
    good = sd > 1e-12
    return (X[:, good] - mu[good]) / sd[good]


def residualize(y: np.ndarray, X: np.ndarray) -> np.ndarray:
    A = np.column_stack([np.ones(len(y)), X])
    beta = np.linalg.lstsq(A, y, rcond=None)[0]
    return y - A @ beta


def fisher_mean(rs) -> float:
    v = np.asarray([r for r in rs if np.isfinite(r)], float)
    if not len(v):
        return float("nan")
    z = np.arctanh(np.clip(v, -0.999999, 0.999999))
    return float(np.tanh(z.mean()))


def safe_rho(a: np.ndarray, b: np.ndarray) -> float:
    r = float(spearmanr(a, b).statistic)
    if not np.isfinite(r):
        raise RuntimeError("non-finite Spearman RSA")
    return r


def latest_completed_adapter(root: Path) -> Path:
    candidates = []
    if root.exists():
        for d in root.iterdir():
            if d.is_dir() and (d / "summary.json").exists() and (d / "adapter").is_dir():
                candidates.append(d)
    if not candidates:
        raise FileNotFoundError(f"No completed adapter under {root}")
    return sorted(candidates)[-1] / "adapter"


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    return (hidden * mask).sum(dim=1) / mask.sum(dim=1).clamp_min(1.0)


def encode_texts(model, tokenizer, texts: list[str], device: str, batch_size: int = 64) -> np.ndarray:
    import torch

    all_vec = []
    for start in range(0, len(texts), batch_size):
        batch = [PREFIX + t for t in texts[start : start + batch_size]]
        enc = tokenizer(batch, padding=True, truncation=True, max_length=128, return_tensors="pt")
        attention = enc["attention_mask"].to(device)
        enc = {k: v.to(device) for k, v in enc.items()}
        with torch.inference_mode():
            out = model(**enc, return_dict=True)
            pooled = masked_mean(out.last_hidden_state, attention.bool())
            pooled = torch.nn.functional.normalize(pooled, p=2, dim=1)
        all_vec.append(pooled.cpu().numpy().astype(np.float64))
    return np.concatenate(all_vec, axis=0)


def load_base(device: str):
    from transformers import AutoModel, AutoTokenizer

    tok = AutoTokenizer.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    model = AutoModel.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    model.eval().to(device)
    return tok, model


def load_adapter(adapter: Path, device: str):
    from peft import PeftModel
    from transformers import AutoModel, AutoTokenizer

    if not adapter.is_dir():
        raise FileNotFoundError(adapter)
    tok = AutoTokenizer.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    base = AutoModel.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
    model = PeftModel.from_pretrained(base, adapter)
    model.eval().to(device)
    return tok, model


def build_model_rdms(blocks: dict[str, list[str]], device: str):
    import torch

    adapter_010 = latest_completed_adapter(LAMBDA_010_ROOT)
    specs = {
        "base": None,
        "lambda_0": TEXT_ONLY_ADAPTER,
        "lambda_0p10": adapter_010,
        "lambda_1": LAMBDA_1_ADAPTER,
    }
    flat_texts = [t for s in SESSIONS for t in blocks[s]]
    out = {}
    provenance = {}
    for label, adapter in specs.items():
        print(f"Loading model arm: {label}", flush=True)
        if adapter is None:
            tok, model = load_base(device)
        else:
            tok, model = load_adapter(adapter, device)
        emb = encode_texts(model, tok, flat_texts, device)
        if emb.shape[0] != 400:
            raise RuntimeError(f"expected 400 sentence embeddings, got {emb.shape}")
        out[label] = {}
        offset = 0
        for s in SESSIONS:
            e = emb[offset : offset + 50]
            out[label][s] = squareform(pdist(e, metric="cosine"))
            offset += 50
        provenance[label] = None if adapter is None else str(adapter.resolve())
        del model
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
    return out, provenance


def bootstrap_ci(values: np.ndarray, seed: int = 20260825, nboot: int = 10000):
    values = np.asarray(values, float)
    rng = np.random.default_rng(seed)
    means = np.empty(nboot)
    for i in range(nboot):
        means[i] = rng.choice(values, size=len(values), replace=True).mean()
    return [float(np.quantile(means, 0.025)), float(np.quantile(means, 0.975))]


def signflip_mc(values: np.ndarray, seed: int = 20260825, nperm: int = 200000, batch: int = 10000):
    values = np.asarray(values, float)
    observed = float(values.mean())
    rng = np.random.default_rng(seed)
    ge = 0
    done = 0
    while done < nperm:
        n = min(batch, nperm - done)
        signs = rng.choice(np.array([-1.0, 1.0]), size=(n, len(values)), replace=True)
        means = (signs * values[None, :]).mean(axis=1)
        ge += int(np.sum(means >= observed - 1e-15))
        done += n
    return {
        "n": int(len(values)),
        "observed_mean": observed,
        "n_permutations": int(nperm),
        "seed": int(seed),
        "p_one_sided_ge": float((ge + 1) / (nperm + 1)),
        "plus_one_correction": True,
    }


def write_csv(path: Path, rows: list[dict]):
    if not rows:
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-root", type=Path, default=Path("data/raw/tmnred"))
    ap.add_argument("--input-freeze", type=Path, default=Path("outputs/tmnred_representation_input_materialization/latest/summary.json"))
    ap.add_argument("--output-dir", type=Path, default=Path("outputs/tmnred_e5_transfer_v1/latest"))
    ap.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    args = ap.parse_args()

    import torch

    device = "cuda" if args.device == "auto" and torch.cuda.is_available() else args.device
    if args.device == "auto" and not torch.cuda.is_available():
        device = "cpu"
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("CUDA requested but unavailable")

    root = args.data_root.resolve()
    outdir = args.output_dir.resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    freeze = json.loads(args.input_freeze.read_text())
    if freeze.get("ready_subjects_all_8_sessions") != READY_SUBJECTS:
        raise SystemExit("frozen TMNRED subject cohort mismatch")
    if freeze.get("excluded_subjects") != ["sub-25"]:
        raise SystemExit("unexpected frozen exclusion list")
    if freeze.get("item_cohort_failures"):
        raise SystemExit("TMNRED item cohort freeze is not clean")
    for s in SESSIONS:
        if freeze["item_coverage_by_session"][s]["n_core_items"] != 50:
            raise SystemExit(f"unexpected frozen core size for {s}")

    blocks = stimulus_blocks(root / "derivatives/source material/source material.xlsx")
    model_rdms, model_provenance = build_model_rdms(blocks, device)

    session_rows = []
    by_subject = {sub: {label: {"raw": [], "resid": []} for label in model_rdms} for sub in READY_SUBJECTS}

    for subject in READY_SUBJECTS:
        for session in SESSIONS:
            arr, emap = load_signal(root, subject, session)
            items = sorted(emap)
            if len(items) < 30:
                raise RuntimeError(f"{subject}/{session}: retained item count below frozen QC")
            epidx = [emap[i] - 1 for i in items]
            feat = zscore_cols(row_mean_features(arr)[epidx, :])
            neural = pdist(feat, metric="correlation")
            X = nuisance_for_items(items, blocks[session])
            neural_resid = residualize(neural, X)

            row = {"subject": subject, "session": session, "n_items": len(items), "n_edges": len(neural)}
            ix = np.asarray(items, dtype=int) - 1
            for label in ["base", "lambda_0", "lambda_0p10", "lambda_1"]:
                model_square = model_rdms[label][session]
                model_vec = squareform(model_square[np.ix_(ix, ix)], checks=False)
                if len(model_vec) != len(neural):
                    raise RuntimeError("model/neural edge count mismatch")
                raw = safe_rho(neural, model_vec)
                resid = safe_rho(neural_resid, residualize(model_vec, X))
                row[f"raw_{label}"] = raw
                row[f"resid_{label}"] = resid
                by_subject[subject][label]["raw"].append(raw)
                by_subject[subject][label]["resid"].append(resid)
            row["delta_resid_0p10_vs_0"] = row["resid_lambda_0p10"] - row["resid_lambda_0"]
            row["delta_raw_0p10_vs_0"] = row["raw_lambda_0p10"] - row["raw_lambda_0"]
            session_rows.append(row)

    subject_rows = []
    for subject in READY_SUBJECTS:
        row = {"subject": subject}
        for label in ["base", "lambda_0", "lambda_0p10", "lambda_1"]:
            row[f"raw_{label}"] = fisher_mean(by_subject[subject][label]["raw"])
            row[f"resid_{label}"] = fisher_mean(by_subject[subject][label]["resid"])
        row["delta_resid_0p10_vs_0"] = row["resid_lambda_0p10"] - row["resid_lambda_0"]
        row["delta_raw_0p10_vs_0"] = row["raw_lambda_0p10"] - row["raw_lambda_0"]
        row["delta_resid_1_vs_0"] = row["resid_lambda_1"] - row["resid_lambda_0"]
        row["delta_raw_1_vs_0"] = row["raw_lambda_1"] - row["raw_lambda_0"]
        subject_rows.append(row)

    primary_delta = np.asarray([r["delta_resid_0p10_vs_0"] for r in subject_rows], float)
    raw_delta = np.asarray([r["delta_raw_0p10_vs_0"] for r in subject_rows], float)
    secondary_delta_1 = np.asarray([r["delta_resid_1_vs_0"] for r in subject_rows], float)

    def arm_summary(metric_prefix: str):
        return {
            label: float(np.mean([r[f"{metric_prefix}_{label}"] for r in subject_rows]))
            for label in ["base", "lambda_0", "lambda_0p10", "lambda_1"]
        }

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "protocol": "docs/tmnred_e5_transfer_protocol_v1.md",
        "analysis_status": "prospective frozen external model transfer",
        "dataset": "TMNRED",
        "openneuro_accession": "ds005383",
        "published_snapshot": "1.0.0",
        "model_id": MODEL_ID,
        "model_revision": MODEL_REVISION,
        "model_provenance": model_provenance,
        "device": device,
        "n_subjects": len(READY_SUBJECTS),
        "subjects": READY_SUBJECTS,
        "excluded_subject": "sub-25",
        "resampled_subject": "sub-23",
        "eeg_primary_representation": "row_mean_all",
        "eeg_window_seconds": [0.0, 2.0],
        "model_text": "original Chinese sentence with query: prefix",
        "model_rdm": "cosine distance of L2-normalized attention-mask mean final-hidden-state embeddings",
        "neural_rdm": "feature-wise z-score across retained items, correlation distance",
        "primary_rsa": "Spearman correlation after separately residualizing neural and model edge vectors against the four frozen nuisance RDMs",
        "participant_aggregation": "Fisher-z mean across eight within-session RSA values",
        "nuisance_rdms": [
            "absolute trial-position difference",
            "CJK character-count difference",
            "punctuation-count difference",
            "CJK character-set Jaccard distance",
        ],
        "mean_participant_resid_rsa_by_arm": arm_summary("resid"),
        "mean_participant_raw_rsa_by_arm": arm_summary("raw"),
        "primary_contrast": {
            "name": "lambda_0p10_minus_lambda_0_residual_RSA",
            "mean_delta": float(primary_delta.mean()),
            "median_delta": float(np.median(primary_delta)),
            "fraction_positive": float(np.mean(primary_delta > 0)),
            "bootstrap_95ci_mean": bootstrap_ci(primary_delta),
            "one_sided_signflip": signflip_mc(primary_delta),
        },
        "secondary_raw_contrast": {
            "mean_delta": float(raw_delta.mean()),
            "median_delta": float(np.median(raw_delta)),
            "fraction_positive": float(np.mean(raw_delta > 0)),
            "bootstrap_95ci_mean": bootstrap_ci(raw_delta, seed=20260826),
            "one_sided_signflip": signflip_mc(raw_delta, seed=20260826),
        },
        "secondary_lambda_1_resid_contrast": {
            "mean_delta": float(secondary_delta_1.mean()),
            "median_delta": float(np.median(secondary_delta_1)),
            "fraction_positive": float(np.mean(secondary_delta_1 > 0)),
        },
        "guardrails": [
            "TMNRED is not used for model training or tuning.",
            "The primary EEG target is the prospectively frozen all-sensor temporal mean only.",
            "Amplitude-SD and 8-bin TMNRED representations cannot define primary transfer success.",
            "No model arm, time window, layer, pooling rule, nuisance set, or lambda is selected from TMNRED transfer outcomes.",
        ],
    }

    write_csv(outdir / "session_results.csv", session_rows)
    write_csv(outdir / "subject_results.csv", subject_rows)
    (outdir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": "ok", "primary_contrast": summary["primary_contrast"], "output_dir": str(outdir)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
