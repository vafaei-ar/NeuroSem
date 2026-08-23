#!/usr/bin/env python3
"""Evaluate the frozen four-arm NeuroSem BERT tuning experiment on run 07.

This script is for evaluation only. It never updates model parameters. It loads the
frozen pretrained BERT model and the latest saved LoRA adapters for text_only, neural,
and shuffled_neural, generates run-07 embeddings with the locked representation rule,
and runs the existing locked nuisance-controlled run-07 RSA for each arm.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

ARMS = ["base", "text_only", "neural", "shuffled_neural"]
MODEL_ID = "google-bert/bert-base-chinese"
MODEL_REVISION = "8d2a91f91cc38c96bb8b4556ba70c392f8d5ee55"
SUBJECTS = [
    "sub-04", "sub-05", "sub-06", "sub-07", "sub-08",
    "sub-09", "sub-10", "sub-13", "sub-14", "sub-15",
]


def latest_dir(root: Path, required: str) -> Path:
    candidates = [d for d in root.iterdir() if d.is_dir() and (d / required).exists()] if root.exists() else []
    if not candidates:
        raise FileNotFoundError(f"No directory containing {required} under {root}")
    return sorted(candidates)[-1]


def read_rows(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[str] = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            rows.append(str(value))
    return rows


def masked_mean(hidden, mask):
    mask = mask.to(hidden.dtype).unsqueeze(-1)
    return (hidden * mask).sum(dim=1) / mask.sum(dim=1).clamp_min(1.0)


def generate_embeddings(model, tokenizer, texts, device: str, batch_size: int, max_length: int) -> np.ndarray:
    import torch

    chunks = []
    model.eval()
    with torch.inference_mode():
        for start in range(0, len(texts), batch_size):
            batch = texts[start:start + batch_size]
            encoded = tokenizer(
                batch,
                padding=True,
                truncation=True,
                max_length=max_length,
                return_tensors="pt",
                return_special_tokens_mask=True,
            )
            special = encoded.pop("special_tokens_mask")
            attention = encoded["attention_mask"]
            token_mask = attention.bool() & ~special.bool()
            encoded = {k: v.to(device) for k, v in encoded.items()}
            outputs = model(**encoded, output_hidden_states=True, return_dict=True)
            pooled = masked_mean(outputs.hidden_states[-1], token_mask.to(device))
            chunks.append(pooled.cpu().numpy().astype(np.float32))
    return np.concatenate(chunks, axis=0)


def run(cmd: list[str]) -> None:
    print("+", " ".join(cmd), flush=True)
    subprocess.run(cmd, check=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Evaluate all frozen BERT tuning arms on run-07 neural holdout.")
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--tuning-root", type=Path, default=Path("outputs/bert_neural_tuning_v1"))
    parser.add_argument("--feature-root", type=Path, default=Path("outputs/chineseeeg_row_features/run-07"))
    parser.add_argument("--embedding-output", type=Path, default=Path("outputs/bert_neurosem_run07_embeddings_v1"))
    parser.add_argument("--rsa-output", type=Path, default=Path("outputs/bert_neurosem_run07_rsa_v1"))
    parser.add_argument("--batch-size", type=int, default=128)
    parser.add_argument("--max-length", type=int, default=64)
    parser.add_argument("--permutations", type=int, default=10000)
    parser.add_argument("--workers", type=int, default=16)
    parser.add_argument("--chunk-size", type=int, default=50)
    parser.add_argument("--device", choices=["auto", "cpu", "cuda"], default="auto")
    args = parser.parse_args()

    import torch
    from peft import PeftModel
    from transformers import AutoModelForMaskedLM, AutoTokenizer

    dataset = args.dataset.expanduser().resolve()
    workbook = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / "segmented_Chinense_novel_run_7.xlsx"
    if not workbook.exists():
        raise SystemExit(f"Run-07 workbook not materialized: {workbook}")
    texts = read_rows(workbook)
    if not texts:
        raise SystemExit("No run-07 text rows found")

    if args.device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    else:
        device = args.device
    if device == "cuda" and not torch.cuda.is_available():
        raise SystemExit("CUDA requested but unavailable")

    tokenizer = AutoTokenizer.from_pretrained(MODEL_ID, revision=MODEL_REVISION)

    for arm in ARMS:
        print(f"\n=== {arm} | run-07 final neural holdout ===", flush=True)
        base = AutoModelForMaskedLM.from_pretrained(MODEL_ID, revision=MODEL_REVISION)
        adapter_dir = None
        if arm != "base":
            tuned = latest_dir(args.tuning_root / arm, "summary.json")
            adapter_dir = tuned / "adapter"
            if not adapter_dir.exists():
                raise SystemExit(f"Missing saved adapter for {arm}: {adapter_dir}")
            model = PeftModel.from_pretrained(base, adapter_dir)
        else:
            model = base
        model.to(device)

        emb = generate_embeddings(model, tokenizer, texts, device, args.batch_size, args.max_length)
        if emb.shape[0] != len(texts) or not np.isfinite(emb).all():
            raise RuntimeError(f"Invalid embeddings for {arm}")

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        emb_dir = (args.embedding_output / arm / "run-07" / stamp).resolve()
        emb_dir.mkdir(parents=True, exist_ok=False)
        np.save(emb_dir / "embeddings.npy", emb)
        (emb_dir / "texts.json").write_text(json.dumps(texts, ensure_ascii=False, indent=2), encoding="utf-8")
        summary = {
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "arm": arm,
            "model_id": MODEL_ID,
            "model_revision": MODEL_REVISION,
            "adapter_dir": None if adapter_dir is None else str(adapter_dir.resolve()),
            "run_number": 7,
            "n_rows": len(texts),
            "embedding_shape": list(emb.shape),
            "pooling": "final hidden layer mean over non-special, non-padding tokens",
            "max_length": args.max_length,
            "device": device,
            "evaluation_only": True,
        }
        (emb_dir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        print(f"Embedding output: {emb_dir}")

        run([
            sys.executable,
            "scripts/analysis/assess_chineseeeg_run07_holdout_fast.py",
            "--embedding-root", str(args.embedding_output / arm / "run-07"),
            "--feature-root", str(args.feature_root),
            "--subjects", *SUBJECTS,
            "--permutations", str(args.permutations),
            "--workers", str(args.workers),
            "--chunk-size", str(args.chunk_size),
            "--output-dir", str(args.rsa_output / arm / "run-07"),
        ])

        del model, base
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    print("\nFrozen four-arm run-07 evaluation complete.")
    print("No parameter updates were performed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
