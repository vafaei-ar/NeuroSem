#!/usr/bin/env python3
"""Validate ChineseEEG run-level mapping between text rows, BIDS event markers, and author embeddings.

The authors' XLSX files contain chapter-number rows mixed with displayed text rows. This
script therefore reports both the raw workbook-row count and a conservative text-only count
instead of silently assuming that the first row is the only chapter marker.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import mne
import numpy as np
import openpyxl


def read_tsv(path: Path):
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def chapter_number(value):
    """Return an integer chapter number for simple numeric chapter rows, else None."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if re.fullmatch(r"\d+", s):
            return int(s)
    return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--subject", default="sub-04")
    parser.add_argument("--session", default="ses-LittlePrince")
    parser.add_argument("--run", default="run-01")
    parser.add_argument("--run-number", type=int, default=1)
    parser.add_argument("--derivative", default="filtered_0.5_30")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/chineseeeg_text_alignment"))
    args = parser.parse_args()

    dataset = args.dataset.expanduser().resolve()
    base = dataset / "derivatives" / args.derivative / args.subject / args.session / "eeg" / f"{args.subject}_{args.session}_task-reading_{args.run}"
    vhdr = Path(str(base) + "_eeg.vhdr")
    events_tsv = Path(str(base) + "_events.tsv")
    text_xlsx = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / f"segmented_Chinense_novel_run_{args.run_number}.xlsx"
    embed_npy = dataset / "derivatives" / "text_embeddings" / "LittlePrince_text_embedding" / f"text_embedding_run_{args.run_number}.npy"

    required = [vhdr, events_tsv, text_xlsx, embed_npy]
    missing = [str(p) for p in required if not p.exists()]
    if missing:
        raise SystemExit("Missing required files:\n" + "\n".join(missing))

    wb = openpyxl.load_workbook(text_xlsx, read_only=True, data_only=True)
    ws = wb.active
    workbook_rows = []
    for i in range(2, ws.max_row + 1):
        value = ws.cell(row=i, column=1).value
        if value is not None:
            workbook_rows.append({"xlsx_row": i, "value": value, "chapter": chapter_number(value)})
    if not workbook_rows:
        raise SystemExit("No non-empty rows parsed from xlsx")

    chapter_rows = [r for r in workbook_rows if r["chapter"] is not None]
    text_rows = [r for r in workbook_rows if r["chapter"] is None]
    if not chapter_rows:
        raise SystemExit("No numeric chapter row found in XLSX; inspect workbook structure")

    start_chapter = int(chapter_rows[0]["chapter"])
    start_marker = f"CH{start_chapter:02d}"

    raw = mne.io.read_raw_brainvision(vhdr, preload=False, verbose="ERROR")
    annotations = [(float(o), float(d), str(desc)) for o, d, desc in zip(raw.annotations.onset, raw.annotations.duration, raw.annotations.description)]
    descriptions = [x[2] for x in annotations]

    event_rows = read_tsv(events_tsv)
    event_trial_types = [str(r.get("trial_type", "")) for r in event_rows]
    start_indices = [i for i, r in enumerate(event_rows) if str(r.get("trial_type", "")) == start_marker]
    if not start_indices:
        raise SystemExit(f"Start chapter marker not found in BIDS events.tsv: {start_marker}")
    start_idx = start_indices[0]
    relevant = event_rows[start_idx:]

    segments = []
    malformed = []
    open_start = None
    for i, row in enumerate(relevant, start=start_idx):
        tt = str(row.get("trial_type", ""))
        try:
            onset = float(row.get("onset", "nan"))
        except Exception:
            continue
        if tt == "ROWS":
            if open_start is not None:
                malformed.append({"event_index": i, "type": "ROWS_before_previous_ROWE", "onset": onset})
            open_start = onset
        elif tt == "ROWE":
            if open_start is None:
                malformed.append({"event_index": i, "type": "ROWE_without_ROWS", "onset": onset})
            else:
                segments.append((open_start, onset))
                open_start = None
    if open_start is not None:
        malformed.append({"type": "unclosed_ROWS", "onset": open_start})

    embeddings = np.load(embed_npy, mmap_mode="r")
    n_segments = len(segments)
    n_workbook_rows = len(workbook_rows)
    n_text_rows = len(text_rows)
    n_chapter_rows = len(chapter_rows)
    n_embeddings = int(embeddings.shape[0]) if embeddings.ndim >= 1 else 0
    durations = np.array([b - a for a, b in segments], dtype=float) if segments else np.array([])

    bids_chapter_markers = [tt for tt in event_trial_types if re.fullmatch(r"CH\d+", tt)]
    xlsx_chapters = [int(r["chapter"]) for r in chapter_rows]
    bids_chapters = [int(tt[2:]) for tt in bids_chapter_markers]

    # The author embedding script embeds every non-empty XLSX row from row 2 onward,
    # including numeric chapter rows. Test that behavior explicitly.
    checks = {
        "start_marker_present_in_bids_events": start_marker in event_trial_types,
        "workbook_rows_equal_embeddings": n_workbook_rows == n_embeddings,
        "text_only_rows_equal_segments": n_text_rows == n_segments,
        "segments_equal_embeddings": n_segments == n_embeddings,
        "xlsx_chapter_count": n_chapter_rows,
        "bids_chapter_marker_count": len(bids_chapter_markers),
        "xlsx_chapter_sequence_equals_bids": xlsx_chapters == bids_chapters,
        "all_segments_positive_duration": bool(np.all(durations > 0)) if durations.size else False,
        "no_malformed_ROWS_ROWE_sequence": len(malformed) == 0,
        "mne_annotation_count_equals_bids_event_count": len(annotations) == len(event_rows),
    }

    # Pair sequentially only for diagnostic display. Do not claim semantic alignment unless
    # the appropriate count check passes.
    diagnostic_pairs = []
    candidate_values = [r["value"] for r in text_rows] if n_text_rows == n_segments else [r["value"] for r in workbook_rows]
    for i in range(min(10, len(candidate_values), n_segments)):
        diagnostic_pairs.append({
            "index": i,
            "text": str(candidate_values[i]),
            "start_sec": segments[i][0],
            "end_sec": segments[i][1],
            "duration_sec": segments[i][1] - segments[i][0],
        })

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "pilot": {"subject": args.subject, "session": args.session, "run": args.run, "derivative": args.derivative},
        "text_file": str(text_xlsx),
        "embedding_file": str(embed_npy),
        "start_chapter": start_chapter,
        "start_marker": start_marker,
        "workbook_nonempty_rows_after_header": n_workbook_rows,
        "xlsx_numeric_chapter_rows": [{"xlsx_row": r["xlsx_row"], "chapter": r["chapter"]} for r in chapter_rows],
        "text_only_rows": n_text_rows,
        "paired_ROWS_ROWE_segments": n_segments,
        "embedding_shape": list(embeddings.shape),
        "event_tsv_rows": len(event_rows),
        "bids_chapter_markers": bids_chapter_markers,
        "annotation_counts": dict(Counter(descriptions)),
        "malformed_rows_sequence": malformed,
        "segment_duration_sec": {
            "min": float(durations.min()) if durations.size else None,
            "median": float(np.median(durations)) if durations.size else None,
            "max": float(durations.max()) if durations.size else None,
        },
        "checks": checks,
        "first_diagnostic_pairs": diagnostic_pairs,
    }

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = (args.output_dir / stamp).resolve()
    out.mkdir(parents=True, exist_ok=False)
    (out / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Validation output: {out}")
    print(f"Start chapter/marker: {start_chapter} / {start_marker}")
    print(f"Workbook non-empty rows after header: {n_workbook_rows}")
    print(f"Numeric chapter rows: {n_chapter_rows} -> {xlsx_chapters}")
    print(f"Text-only rows: {n_text_rows}")
    print(f"ROWS-ROWE segments: {n_segments}")
    print(f"Embedding shape: {tuple(embeddings.shape)}")
    print(f"BIDS chapter markers: {bids_chapter_markers}")
    print(f"BIDS events: {len(event_rows)} | MNE annotations: {len(annotations)}")
    print(f"Malformed ROWS/ROWE sequences: {len(malformed)}")
    print("Checks:")
    for key, value in checks.items():
        print(f"  {key}: {value}")
    print("First diagnostic pairs:")
    for item in diagnostic_pairs[:5]:
        print(f"  {item['index']:03d} {item['start_sec']:.3f}-{item['end_sec']:.3f}s  {item['text']}")
    print("NOTE: diagnostic pairs are not accepted as semantic alignment unless the relevant count check passes.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
