#!/usr/bin/env python3
"""Build the first verified ChineseEEG row-level representation manifest.

The manifest preserves the canonical one-to-one ordering among workbook rows,
BIDS ROWS/ROWE segments, and author embedding indices. Numeric chapter rows are
kept for index integrity but flagged as non-semantic structural rows.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl


def read_events(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def is_numeric_chapter(value: object) -> tuple[bool, int | None]:
    if isinstance(value, bool):
        return False, None
    if isinstance(value, (int, np.integer)):
        return True, int(value)
    if isinstance(value, float) and float(value).is_integer():
        return True, int(value)
    text = str(value).strip()
    if text.isdigit():
        return True, int(text)
    return False, None


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a verified ChineseEEG row-level pilot manifest.")
    parser.add_argument("dataset", type=Path, nargs="?", default=Path("data/raw/chineseeeg"))
    parser.add_argument("--subject", default="sub-04")
    parser.add_argument("--session", default="ses-LittlePrince")
    parser.add_argument("--run", default="run-01")
    parser.add_argument("--run-number", type=int, default=1)
    parser.add_argument("--derivative", default="filtered_0.5_30")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/chineseeeg_pilot_manifest"))
    args = parser.parse_args()

    dataset = args.dataset.expanduser().resolve()
    base = dataset / "derivatives" / args.derivative / args.subject / args.session / "eeg" / f"{args.subject}_{args.session}_task-reading_{args.run}"
    events_tsv = Path(str(base) + "_events.tsv")
    text_xlsx = dataset / "derivatives" / "novels" / "segmented_novel" / "LittlePrince" / f"segmented_Chinense_novel_run_{args.run_number}.xlsx"
    embed_npy = dataset / "derivatives" / "text_embeddings" / "LittlePrince_text_embedding" / f"text_embedding_run_{args.run_number}.npy"

    required = [events_tsv, text_xlsx, embed_npy]
    missing = [str(p) for p in required if not p.exists()]
    if missing:
        raise SystemExit("Missing required files:\n" + "\n".join(missing))

    wb = openpyxl.load_workbook(text_xlsx, read_only=True, data_only=True)
    ws = wb.active
    workbook_rows = []
    for excel_row in range(2, ws.max_row + 1):
        value = ws.cell(row=excel_row, column=1).value
        if value is not None:
            workbook_rows.append((excel_row, value))

    events = read_events(events_tsv)
    chapter_context = None
    row_segments: list[dict[str, object]] = []
    open_row: tuple[float, int, str | None] | None = None
    malformed = 0

    for event_index, event in enumerate(events):
        trial_type = (event.get("trial_type") or "").strip()
        onset = float(event["onset"])
        if trial_type.startswith("CH") and len(trial_type) >= 4 and trial_type[2:].isdigit():
            chapter_context = int(trial_type[2:])
        elif trial_type == "ROWS":
            if open_row is not None:
                malformed += 1
            open_row = (onset, event_index, f"CH{chapter_context:02d}" if chapter_context is not None else None)
        elif trial_type == "ROWE":
            if open_row is None:
                malformed += 1
                continue
            start_onset, start_event_index, chapter_marker = open_row
            row_segments.append({
                "start_sec": start_onset,
                "end_sec": onset,
                "duration_sec": onset - start_onset,
                "rows_event_index": start_event_index,
                "rowe_event_index": event_index,
                "chapter_marker": chapter_marker,
            })
            open_row = None

    if open_row is not None:
        malformed += 1

    embeddings = np.load(embed_npy, mmap_mode="r")
    n_embeddings = int(embeddings.shape[0]) if embeddings.ndim >= 1 else 0

    if malformed:
        raise SystemExit(f"Malformed ROWS/ROWE sequence count: {malformed}")
    if not (len(workbook_rows) == len(row_segments) == n_embeddings):
        raise SystemExit(
            "Count mismatch: "
            f"workbook={len(workbook_rows)}, segments={len(row_segments)}, embeddings={n_embeddings}"
        )

    out_rows: list[dict[str, object]] = []
    chapter_rows = 0
    for index, ((excel_row, value), segment) in enumerate(zip(workbook_rows, row_segments)):
        is_chapter, chapter_number = is_numeric_chapter(value)
        chapter_rows += int(is_chapter)
        expected_marker = f"CH{chapter_number:02d}" if chapter_number is not None else None
        chapter_match = None if not is_chapter else expected_marker == segment["chapter_marker"]
        text = str(value)
        out_rows.append({
            "alignment_index": index,
            "excel_row": excel_row,
            "embedding_index": index,
            "text": text,
            "is_chapter_row": is_chapter,
            "chapter_number": chapter_number if chapter_number is not None else "",
            "chapter_marker_context": segment["chapter_marker"] or "",
            "chapter_row_marker_match": "" if chapter_match is None else chapter_match,
            "semantic_eligible": not is_chapter,
            "start_sec": segment["start_sec"],
            "end_sec": segment["end_sec"],
            "duration_sec": segment["duration_sec"],
            "rows_event_index": segment["rows_event_index"],
            "rowe_event_index": segment["rowe_event_index"],
            "run_position_fraction": index / max(1, len(workbook_rows) - 1),
            "char_count": len(text),
        })

    chapter_mismatches = [r for r in out_rows if r["is_chapter_row"] and r["chapter_row_marker_match"] is not True]
    if chapter_mismatches:
        raise SystemExit(f"Chapter-row marker mismatch for {len(chapter_mismatches)} rows")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = (args.output_dir / stamp).resolve()
    out.mkdir(parents=True, exist_ok=False)
    csv_path = out / "pilot_manifest.csv"
    fieldnames = list(out_rows[0].keys())
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    report = [
        "# ChineseEEG pilot representation manifest",
        "",
        f"- Subject/session/run: `{args.subject} / {args.session} / {args.run}`",
        f"- Canonical aligned rows: **{len(out_rows)}**",
        f"- Semantic-eligible text rows: **{len(out_rows) - chapter_rows}**",
        f"- Structural chapter rows: **{chapter_rows}**",
        f"- Author embedding shape: **{tuple(embeddings.shape)}**",
        "- Alignment invariant: workbook row index = EEG ROWS/ROWE segment index = author embedding index.",
        "- Chapter rows are preserved for index integrity and excluded from semantic analyses.",
        "",
        "## First rows",
        "",
    ]
    for row in out_rows[:10]:
        report.append(
            f"- {row['alignment_index']:03d} | {row['start_sec']:.3f}-{row['end_sec']:.3f}s | "
            f"chapter={row['chapter_marker_context']} | semantic={row['semantic_eligible']} | {row['text']}"
        )
    (out / "report.md").write_text("\n".join(report) + "\n", encoding="utf-8")

    print(f"Manifest output: {out}")
    print(f"Aligned rows: {len(out_rows)}")
    print(f"Semantic-eligible rows: {len(out_rows) - chapter_rows}")
    print(f"Structural chapter rows: {chapter_rows}")
    print(f"Embedding shape: {tuple(embeddings.shape)}")
    print("Chapter-row marker checks: PASS")
    print(f"CSV: {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
