#!/usr/bin/env python3
"""Model-blind Garnett Dream segmented-XLSX row-mapping probe.

The ChineseEEG authors document that the non-"display" per-run segmented XLSX
files contain the textual units/lines used for EEG analysis, while companion
"display" XLSX files are transformed for PsychoPy presentation. This probe
materializes only tracked XLSX files under derivatives/novels, inventories
their structure without exporting text, and tests whether exactly one
Garnett-Dream non-display run workbook matches the already-frozen ROWS->ROWE
item count for each chapter/run 1..18.

No EEG sample, model embedding, adapter, RSA, or reliability quantity is loaded.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook

RUN_RE = re.compile(r"run[_-]?(\d{1,2})", re.I)
GARNETT_KEYS = ("garnett", "dream", "langwang", "lang_wang", "lang-wang")


def git_lines(root: Path, *args: str) -> list[str]:
    cp = subprocess.run(["git", "-C", str(root), *args], capture_output=True, text=True, check=False)
    if cp.returncode != 0:
        raise RuntimeError(cp.stderr.strip() or f"git {' '.join(args)} failed")
    return [x for x in cp.stdout.splitlines() if x.strip()]


def annex_get(root: Path, rel: str) -> dict:
    p = root / rel
    before = p.exists() and p.is_file() and p.stat().st_size > 0
    cp = None
    if not before:
        cp = subprocess.run(["git", "-C", str(root), "annex", "get", "--", rel], capture_output=True, text=True, check=False)
    after = p.exists() and p.is_file() and p.stat().st_size > 0
    return {
        "materialized_before": before,
        "materialized_after": after,
        "annex_get_returncode": None if cp is None else cp.returncode,
        "annex_get_stderr_tail": "" if cp is None else cp.stderr[-500:],
    }


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def workbook_structure(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    total_nonempty_rows = 0
    total_nonempty_cells = 0
    row_hashes = []
    for ws in wb.worksheets:
        n_rows = 0
        n_cells = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if not vals:
                continue
            n_rows += 1
            n_cells += len(vals)
            # Hash only, never export text.
            joined = "\u241f".join(vals)
            row_hashes.append(hashlib.sha256(joined.encode("utf-8")).hexdigest())
        sheets.append({"title": ws.title, "nonempty_rows": n_rows, "nonempty_cells": n_cells})
        total_nonempty_rows += n_rows
        total_nonempty_cells += n_cells
    wb.close()
    sequence_hash = hashlib.sha256("\n".join(row_hashes).encode("ascii")).hexdigest() if row_hashes else None
    return {
        "n_sheets": len(sheets),
        "sheets": sheets,
        "nonempty_rows": total_nonempty_rows,
        "nonempty_cells": total_nonempty_cells,
        "row_sequence_sha256": sequence_hash,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-root", type=Path, default=Path("data/raw/chineseeeg"))
    ap.add_argument("--input-freeze", type=Path, default=Path("outputs/garnett_dream_input_materialization/latest/summary.json"))
    ap.add_argument("--output-dir", type=Path, default=Path("outputs/garnett_dream_segmented_xlsx_mapping_probe_v1/latest"))
    args = ap.parse_args()

    root = args.data_root.resolve()
    freeze = json.loads(args.input_freeze.read_text(encoding="utf-8"))
    if not freeze.get("freeze_gate", {}).get("ready_for_reliability"):
        raise SystemExit("Garnett input freeze is not structurally ready")
    expected = {int(k): int(v) for k, v in freeze.get("chapter_item_counts", {}).items()}
    if sorted(expected) != list(range(1, 19)):
        raise SystemExit("Expected frozen Garnett chapters 1..18")

    out = args.output_dir.resolve()
    out.mkdir(parents=True, exist_ok=True)

    tracked = [p for p in git_lines(root, "ls-files", "derivatives/novels") if p.lower().endswith(".xlsx")]
    records = []
    failures = []
    for rel in tracked:
        rec = {"path": rel, **annex_get(root, rel)}
        p = root / rel
        low = rel.lower()
        rec["is_display"] = "display" in low
        rec["looks_garnett"] = any(k in low for k in GARNETT_KEYS)
        m = RUN_RE.search(rel)
        rec["run_from_path"] = int(m.group(1)) if m else None
        if rec["materialized_after"]:
            try:
                rec["size_bytes"] = p.stat().st_size
                rec["sha256"] = sha256(p)
                rec.update(workbook_structure(p))
            except Exception as exc:
                rec["workbook_error"] = f"{type(exc).__name__}: {exc}"
                failures.append({"path": rel, "reason": "workbook_read_error", "error": rec["workbook_error"]})
        else:
            failures.append({"path": rel, "reason": "materialization_failed"})
        records.append(rec)

    candidates_by_run = {}
    matches = []
    ambiguous = []
    missing = []
    for run in range(1, 19):
        candidates = [
            r for r in records
            if r.get("materialized_after")
            and r.get("looks_garnett")
            and not r.get("is_display")
            and r.get("run_from_path") == run
            and not r.get("workbook_error")
        ]
        candidates_by_run[str(run)] = [
            {"path": r["path"], "nonempty_rows": r.get("nonempty_rows"), "sha256": r.get("sha256"), "row_sequence_sha256": r.get("row_sequence_sha256")}
            for r in candidates
        ]
        exact = [r for r in candidates if int(r.get("nonempty_rows") or -1) == expected[run]]
        if len(exact) == 1:
            r = exact[0]
            matches.append({
                "run": run,
                "chapter": run,
                "expected_items": expected[run],
                "path": r["path"],
                "nonempty_rows": r["nonempty_rows"],
                "sha256": r["sha256"],
                "row_sequence_sha256": r["row_sequence_sha256"],
            })
        elif len(exact) == 0:
            missing.append({"run": run, "expected_items": expected[run], "candidate_count": len(candidates)})
        else:
            ambiguous.append({"run": run, "expected_items": expected[run], "paths": [r["path"] for r in exact]})

    with (out / "workbook_inventory.csv").open("w", encoding="utf-8", newline="") as f:
        fields = [
            "path", "materialized_before", "materialized_after", "is_display", "looks_garnett", "run_from_path",
            "size_bytes", "sha256", "n_sheets", "nonempty_rows", "nonempty_cells", "row_sequence_sha256", "workbook_error",
        ]
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader(); w.writerows(records)

    exact_complete = len(matches) == 18 and not ambiguous and not missing and not failures
    summary = {
        "schema_version": 1,
        "dataset": "ChineseEEG Garnett Dream",
        "purpose": "model-blind segmented-XLSX presentation-row mapping probe",
        "loads_eeg_samples": False,
        "computes_reliability_or_rsa": False,
        "computes_model_quantities": False,
        "exports_text": False,
        "source_semantics": {
            "upstream_repository": "ncclabsustech/Chinese_reading_task_eeg_processing",
            "documented_rule": "non-display segmented per-run XLSX contains displayed analysis lines; display XLSX is a transformed PsychoPy presentation file",
            "upstream_readme_blob_sha": "a72763308292e61b4fefad9c978fe9ba48c9877b",
        },
        "expected_items_by_chapter": {str(k): v for k, v in sorted(expected.items())},
        "n_tracked_xlsx": len(tracked),
        "n_materialized_xlsx": sum(bool(r.get("materialized_after")) for r in records),
        "n_exact_run_matches": len(matches),
        "exact_run_matches": matches,
        "missing_runs": missing,
        "ambiguous_runs": ambiguous,
        "failures": failures,
        "candidate_paths_by_run": candidates_by_run,
        "freeze_gate": {
            "exact_row_text_mapping_identified": exact_complete,
            "ready_to_freeze_model_validation_text_mapping": exact_complete,
            "mapping_rule": "For run/chapter 1..18, map CHxx_ROWyyyy to row yyyy of the unique non-display Garnett segmented run XLSX whose nonempty-row count equals the frozen ROWS->ROWE item count." if exact_complete else None,
            "reason": "Unique non-display segmented XLSX row counts exactly match all 18 frozen Garnett ROWS->ROWE item counts." if exact_complete else "An exact one-to-one segmented-XLSX mapping is not yet structurally established for all 18 runs.",
        },
        "guardrails": [
            "No EEG signal sample is opened.",
            "No model embedding, adapter, lambda, RSA, or reliability outcome is loaded or computed.",
            "No stimulus text is exported; only paths, counts, file hashes, and row-sequence hashes are stored.",
            "No mapping is selected using Garnett neural outcomes.",
        ],
    }
    (out / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"status": "ok", "n_tracked_xlsx": len(tracked), "n_exact_run_matches": len(matches), "exact_complete": exact_complete, "output_dir": str(out)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
