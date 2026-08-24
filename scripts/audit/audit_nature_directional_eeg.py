#!/usr/bin/env python3
"""Structural audit for the Scientific Data 2026 directional-word EEG dataset.

This audit is intentionally model-blind. It inventories files and EEG/event metadata but
never loads NeuroSem adapters, computes language-model embeddings, or evaluates neural-model RSA.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path

MODIFIED_RUSSIAN = {"sub1", "sub3", "sub5", "sub10"}


def find_dataset_root(extracted: Path) -> Path:
    candidates = []
    for p in [extracted, *[x for x in extracted.rglob("*") if x.is_dir()]]:
        if (p / "preprocessed").is_dir() and (p / "metadata").is_dir():
            candidates.append(p)
    if not candidates:
        raise FileNotFoundError(f"Could not locate dataset root under {extracted}")
    candidates = sorted(candidates, key=lambda p: (len(p.parts), str(p)))
    return candidates[0]


def xlsx_shape(path: Path) -> tuple[int | None, int | None]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return int(ws.max_row), int(ws.max_column)
    except Exception:
        return None, None


def inspect_epochs(path: Path) -> dict:
    import mne

    epochs = mne.read_epochs(path, preload=False, verbose="ERROR")
    return {
        "n_epochs": len(epochs),
        "n_channels": len(epochs.ch_names),
        "n_times": len(epochs.times),
        "sfreq": float(epochs.info["sfreq"]),
        "tmin": float(epochs.tmin),
        "tmax": float(epochs.tmax),
        "event_id": {str(k): int(v) for k, v in epochs.event_id.items()},
        "channel_names": list(epochs.ch_names),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "dataset",
        nargs="?",
        type=Path,
        default=Path("data/raw/nature_directional_eeg/extracted"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/nature_directional_eeg_audit/latest"),
    )
    args = parser.parse_args()

    extracted = args.dataset.expanduser().resolve()
    root = find_dataset_root(extracted)

    rows: list[dict] = []
    preprocessed = root / "preprocessed"
    for language in ("russian", "spanish"):
        lang_dir = preprocessed / language
        if not lang_dir.is_dir():
            continue
        for subject_dir in sorted(p for p in lang_dir.iterdir() if p.is_dir()):
            subject = subject_dir.name
            fif_files = sorted(subject_dir.glob("*.fif"))
            xlsx_files = sorted(subject_dir.glob("*.xlsx"))
            rec = {
                "language": language,
                "subject": subject,
                "modified_marker_protocol": language == "russian" and subject in MODIFIED_RUSSIAN,
                "fif_count": len(fif_files),
                "xlsx_count": len(xlsx_files),
                "fif_path": str(fif_files[0].relative_to(root)) if fif_files else None,
                "xlsx_path": str(xlsx_files[0].relative_to(root)) if xlsx_files else None,
            }
            if fif_files:
                try:
                    rec.update(inspect_epochs(fif_files[0]))
                    rec["epochs_error"] = None
                except Exception as exc:
                    rec["epochs_error"] = f"{type(exc).__name__}: {exc}"
            if xlsx_files:
                nrows, ncols = xlsx_shape(xlsx_files[0])
                rec["xlsx_rows"] = nrows
                rec["xlsx_columns"] = ncols
            rows.append(rec)

    metadata_path = root / "metadata" / "subject_metadata.json"
    metadata = None
    if metadata_path.exists():
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))

    raw_files = sorted((root / "raw").rglob("*.edf")) if (root / "raw").exists() else []
    fif_files_all = sorted(preprocessed.rglob("*.fif")) if preprocessed.exists() else []
    xlsx_files_all = sorted(preprocessed.rglob("*.xlsx")) if preprocessed.exists() else []

    language_counts: dict[str, int] = {}
    standard_counts: dict[str, int] = {}
    for rec in rows:
        language_counts[rec["language"]] = language_counts.get(rec["language"], 0) + 1
        if not rec["modified_marker_protocol"]:
            standard_counts[rec["language"]] = standard_counts.get(rec["language"], 0) + 1

    summary = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "purpose": "Model-blind structural audit before external NeuroSem validation.",
        "publication_doi": "10.1038/s41597-026-07809-9",
        "zenodo_doi": "10.5281/zenodo.20374418",
        "dataset_root": str(root),
        "raw_edf_count": len(raw_files),
        "preprocessed_fif_count": len(fif_files_all),
        "event_xlsx_count": len(xlsx_files_all),
        "subjects_total": len(rows),
        "subjects_by_language": language_counts,
        "standard_protocol_subjects_by_language": standard_counts,
        "modified_russian_subjects_expected": sorted(MODIFIED_RUSSIAN),
        "metadata_present": metadata is not None,
        "metadata_type": type(metadata).__name__ if metadata is not None else None,
        "subject_inventory": rows,
        "model_blind": True,
        "neural_model_rsa_computed": False,
    }

    out = args.output_dir.expanduser().resolve()
    out.mkdir(parents=True, exist_ok=True)
    (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    csv_fields = [
        "language",
        "subject",
        "modified_marker_protocol",
        "fif_count",
        "xlsx_count",
        "n_epochs",
        "n_channels",
        "n_times",
        "sfreq",
        "tmin",
        "tmax",
        "xlsx_rows",
        "xlsx_columns",
        "fif_path",
        "xlsx_path",
        "epochs_error",
    ]
    with (out / "subject_inventory.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        writer.writeheader()
        for rec in rows:
            writer.writerow(rec)

    print(f"Nature directional EEG audit root: {root}")
    print(f"Subjects: {len(rows)} | by language: {language_counts}")
    print(f"Standard protocol by language: {standard_counts}")
    print(f"Files: EDF={len(raw_files)} FIF={len(fif_files_all)} XLSX={len(xlsx_files_all)}")
    print(f"Audit summary: {out / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
