#!/usr/bin/env python3
import argparse
import csv
import json
import subprocess
from pathlib import Path


def annex_get(root: Path, rel: str) -> dict:
    p = root / rel
    tracked = subprocess.run(
        ["git", "-C", str(root), "ls-files", "--error-unmatch", rel],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    ).returncode == 0
    rec = {"path": rel, "tracked": tracked, "materialized_before": p.exists()}
    if not tracked:
        rec["status"] = "not_tracked"
        return rec
    if not p.exists():
        cp = subprocess.run(
            ["git", "-C", str(root), "annex", "get", "--", rel],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        rec["annex_get_returncode"] = cp.returncode
        rec["annex_get_stdout_tail"] = cp.stdout[-1500:]
        rec["annex_get_stderr_tail"] = cp.stderr[-1500:]
    rec["materialized_after"] = p.exists()
    rec["size_bytes"] = p.stat().st_size if p.exists() else None
    rec["status"] = "materialized" if p.exists() else "not_materializable"
    return rec


def clip(v, n=180):
    if v is None:
        return None
    s = str(v)
    return s if len(s) <= n else s[: n - 3] + "..."


def summarize_csv(path: Path) -> dict:
    out = {"path": str(path), "exists": path.exists()}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    out["n_rows_including_header"] = len(rows)
    out["max_columns"] = max((len(r) for r in rows), default=0)
    out["header"] = [clip(x) for x in rows[0]] if rows else []
    out["sample_rows"] = [[clip(x) for x in r] for r in rows[1:8]]
    return out


def summarize_xlsx(path: Path) -> dict:
    out = {"path": str(path), "exists": path.exists()}
    if not path.exists():
        return out
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        out["openpyxl_error"] = repr(exc)
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        sample = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 8:
                break
            sample.append([clip(x) for x in row])
        sheets.append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_rows": sample,
        })
    out["sheets"] = sheets
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-root", default="data/raw/tmnred")
    ap.add_argument("--output-dir", default="outputs/tmnred_stimulus_metadata_probe/latest")
    args = ap.parse_args()

    root = Path(args.data_root)
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    rel_xlsx = "derivatives/source material/source material.xlsx"
    rel_csv = "derivatives/source material/source material_ses.csv"
    materialization = [annex_get(root, rel_xlsx), annex_get(root, rel_csv)]

    payload = {
        "schema_version": 1,
        "dataset": "TMNRED",
        "model_blind": True,
        "signal_loaded": False,
        "purpose": "Resolve public stimulus/session metadata needed to define a prospective cross-subject semantic analysis unit before any EEG-representation reliability or model analysis.",
        "materialization": materialization,
        "xlsx_summary": summarize_xlsx(root / rel_xlsx),
        "csv_summary": summarize_csv(root / rel_csv),
        "notes": [
            "No EEG signal values, neural reliability scores, model embeddings, or neural-model RSA are computed.",
            "Only public stimulus metadata are materialized; no full EEG payload is downloaded by this task.",
            "This probe is intended to determine whether trial labels and sessions map to stable semantic items/conditions before freezing the TMNRED representation benchmark."
        ],
    }
    (outdir / "summary.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    failed = [r for r in materialization if r.get("status") != "materialized"]
    if failed:
        raise SystemExit(f"TMNRED stimulus metadata unavailable: {failed}")
    print(json.dumps({"status": "ok", "output": str(outdir / 'summary.json')}, indent=2))


if __name__ == "__main__":
    main()
